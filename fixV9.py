# -*- coding: utf-8 -*-
"""
Streamlit 단일 파일 버전: 계좌주간 거래 분석 도구
- 업로드한 Excel(xlsx, xls) 파일을 읽어서 계좌주 간 상계 후보를 매칭하고 결과를 다운로드로 제공합니다.
- openpyxl 미설치 환경에서도 앱이 죽지 않도록 안전하게 동작합니다(스타일/서식은 openpyxl 있을 때만 적용).
- 기본적으로 GitHub의 샘플 데이터가 적용되어 있습니다.

필수 패키지 (예시):
    pip install streamlit pandas requests

선택 패키지 (있으면 더 좋음):
    pip install openpyxl            # xlsx 읽기/쓰기 + 스타일 적용에 필요
    pip install "xlrd==1.2.0"      # 구형 .xls 읽기에 필요
"""

import io
import re
import pandas as pd
import numpy as np
import streamlit as st
import requests

# -------------------- openpyxl(선택) 지연 임포트 --------------------
try:
    from openpyxl import load_workbook  # type: ignore
    from openpyxl.styles import PatternFill, Font  # type: ignore
    from openpyxl.utils import get_column_letter  # type: ignore
    OPENPYXL_AVAILABLE = True
except Exception:  # 모듈이 없거나 불러오기 실패해도 앱은 계속 동작
    OPENPYXL_AVAILABLE = False

# ===================== 기본 설정값 (UI에서 변경) =====================
DEFAULT_AMOUNT_TOLERANCE = 100_000   # 금액 허용 오차 (원)
DEFAULT_TIME_UNIT = "year"           # "day" | "month" | "year"
DEFAULT_WINDOW = 10                  # 시간창 크기 (예: 3개월)
DEFAULT_DATA_URL = "https://github.com/serene-gith/Account-transaction-history-analysis/raw/main/계좌내역.xlsx"
# ====================================================================

IN_METHOD_PATTERN = r"(이체|예금|대체)"
EXCLUDE_NOTE_PATTERN = r"(급여|월급|봉급|급여입금|salary)"

# -------------------------- 유틸 함수 --------------------------
def _to_num(x):
    try:
        return float(str(x).replace(",", "").strip())
    except Exception:
        return np.nan


def _read_excel_any(uploaded_file, filename: str) -> pd.DataFrame:
    """업로드한 파일을 확장자/엔진에 맞게 DataFrame으로 읽습니다.
    - .xlsx/.xlsm: openpyxl 권장(없으면 안내 메시지)
    - .xls: xlrd(1.2.0 필요)
    """
    name = (filename or "").lower()
    if name.endswith(".xls") and not name.endswith(".xlsx"):
        # xlrd 2.x는 xls 미지원 → 1.2.0 설치 필요
        try:
            return pd.read_excel(uploaded_file, dtype=str, engine="xlrd")
        except Exception as e:
            raise RuntimeError(
                "*.xls 파일을 읽으려면 'xlrd==1.2.0' 설치가 필요합니다.\n"
                "명령: pip install \"xlrd==1.2.0\"\n"
                f"원인: {e}"
            )
    else:
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError(
                "xlsx/xlsm 파일을 처리하려면 openpyxl이 필요합니다.\n"
                "명령: pip install openpyxl"
            )
        # openpyxl이 있을 때에만 엔진 지정
        return pd.read_excel(uploaded_file, dtype=str, engine="openpyxl")


@st.cache_data
def load_default_data():
    """GitHub에서 기본 데이터를 다운로드합니다."""
    try:
        response = requests.get(DEFAULT_DATA_URL, timeout=10)
        response.raise_for_status()
        return io.BytesIO(response.content)
    except Exception as e:
        st.error(f"기본 데이터를 불러오는 데 실패했습니다: {e}")
        return None


# ---------------------- 원본 로직 이식 ----------------------

def read_data_from_upload(uploaded_file, filename: str) -> pd.DataFrame:
    df = _read_excel_any(uploaded_file, filename).rename(columns={
        "계좌주":"계좌주","계좌번호":"계좌번호","거래일자":"거래일자","거래시각":"거래시각",
        "거래대금(원)":"거래대금","잔액(원)":"잔액","거래방법":"거래방법","비고(거래내용)":"비고"
    })
    for c in ["계좌주","계좌번호","거래방법","비고"]:
        if c in df.columns:
            df[c] = df[c].astype(str).str.strip()
    # 날짜/시각 파싱
    df["거래일자"] = pd.to_datetime(df.get("거래일자"), errors="coerce").dt.date
    t = pd.to_datetime(df.get("거래시각"), format="%H:%M:%S", errors="coerce")
    df["거래시각"] = t.dt.time
    df["일시"] = pd.to_datetime(df["거래일자"].astype(str) + " " + df["거래시각"].astype(str), errors="coerce")
    # 금액/잔액 수치화
    df["거래대금"] = df.get("거래대금", np.nan).apply(_to_num)
    df["잔액"] = df.get("잔액", np.nan).apply(_to_num)
    # 필수 컬럼 결측 제거
    return df.dropna(subset=["계좌주","일시","거래대금"]).reset_index(drop=True)


def parse_counterparty_and_direction(note: str):
    if not isinstance(note, str):
        return None, None
    s = re.sub(r"\s+", "", note)
    m_out = re.search(r"([가-힣]{2,})에게송금", s)
    if m_out:
        return m_out.group(1), "out"
    return None, None


def extract_inter_owner_transactions(df: pd.DataFrame) -> pd.DataFrame:
    parsed = df.get("비고", pd.Series([None] * len(df))).apply(parse_counterparty_and_direction)
    inter = df.copy()
    inter["상대계좌주"] = [p[0] for p in parsed]
    inter["방향"] = [p[1] for p in parsed]
    inter = inter[inter.get("거래방법", "").astype(str).str.contains("이체|대체|예금", na=False)].copy()
    inter = inter[[
        "계좌주","계좌번호","일시","거래일자","거래시각","거래대금","잔액","거래방법","비고","상대계좌주","방향"
    ]].sort_values(["일시","계좌주"]).reset_index(drop=True)
    inter["orig_index"] = inter.index
    inter["제외비고"] = inter["비고"].fillna("").str.contains(EXCLUDE_NOTE_PATTERN, case=False, regex=True)
    # 제외 건은 방향/상대계좌주 비우기
    inter.loc[inter["제외비고"], ["방향","상대계좌주"]] = None
    return inter


def time_within_window(ts1: pd.Timestamp, ts2: pd.Timestamp, time_unit: str, window: int) -> bool:
    if pd.isna(ts1) or pd.isna(ts2):
        return False
    if time_unit == "day":
        return abs(ts2 - ts1) <= pd.Timedelta(days=window)
    elif time_unit == "month":
        return abs((ts2.year - ts1.year)*12 + (ts2.month - ts1.month)) <= window
    elif time_unit == "year":
        return abs(ts2.year - ts1.year) <= window
    return False


def match_out_to_in_and_backfill(inter: pd.DataFrame, amount_tolerance: int, time_unit: str, window: int) -> pd.DataFrame:
    inter = inter.copy()
    inter["상계후보"] = False
    looks_out = inter["비고"].fillna("").str.contains(r"에게\s*송금")
    inter["수신후보"] = inter["거래방법"].str.contains(IN_METHOD_PATTERN, na=False) & (~looks_out)

    for (owner, counter), g in inter.groupby(["계좌주","상대계좌주"], dropna=False):
        if pd.isna(counter):
            continue
        out_g = g[(g["방향"] == "out") & (~g["제외비고"])].sort_values("일시")
        if out_g.empty:
            continue
        in_pool = inter[(inter["계좌주"] == counter) & (inter["수신후보"]) & (~inter["제외비고"])].sort_values("일시").copy()
        if in_pool.empty:
            continue
        out_rows, in_rows = out_g.to_dict("records"), in_pool.to_dict("records")
        i = j = 0
        while i < len(out_rows) and j < len(in_rows):
            ro, ri = out_rows[i], in_rows[j]
            amt_ok = abs(abs(ro["거래대금"]) - abs(ri["거래대금"])) <= amount_tolerance
            time_ok = time_within_window(ro["일시"], ri["일시"], time_unit, window)
            if amt_ok and time_ok:
                inter.loc[inter["orig_index"].isin([ro["orig_index"], ri["orig_index"]]), "상계후보"] = True
                ri_idx = inter.index[inter["orig_index"] == ri["orig_index"]][0]
                if not bool(inter.at[ri_idx, "제외비고"]):
                    if pd.isna(inter.at[ri_idx, "상대계좌주"]) or inter.at[ri_idx, "상대계좌주"] == "":
                        inter.at[ri_idx, "상대계좌주"] = owner
                    if pd.isna(inter.at[ri_idx, "방향"]) or inter.at[ri_idx, "방향"] == "":
                        inter.at[ri_idx, "방향"] = "in"
                i += 1
                j += 1
            else:
                if ro["일시"] <= ri["일시"]:
                    i += 1
                else:
                    j += 1
    return inter


# ----------------------- 엑셀 빌드 & 스타일 -----------------------
if OPENPYXL_AVAILABLE:
    from openpyxl import Workbook  # type: ignore

    def _style_owner_sheet(ws):
        header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        col_amount = header_map.get("거래대금")
        col_balance = header_map.get("잔액")
        col_candidate = header_map.get("상계후보")

        if col_amount:
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=col_amount).number_format = "#,##0"
        if col_balance:
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=col_balance).number_format = "#,##0"

        yellow = PatternFill(start_color="FFF9C0", end_color="FFF9C0", fill_type="solid")
        if col_candidate:
            for r in range(2, ws.max_row + 1):
                val = ws.cell(row=r, column=col_candidate).value
                if str(val).strip().lower() in ("true", "1"):
                    for c in range(1, ws.max_column + 1):
                        ws.cell(row=r, column=c).fill = yellow

        # 합계 행
        last_row = ws.max_row
        total_row = last_row + 1
        ws.cell(row=total_row, column=1, value="합계").font = Font(bold=True)
        if col_amount:
            L = get_column_letter(col_amount)
            ws.cell(row=total_row, column=col_amount, value=f"=SUM({L}2:{L}{last_row})").font = Font(bold=True)
            ws.cell(row=total_row, column=col_amount).number_format = "#,##0"
        if col_balance:
            L = get_column_letter(col_balance)
            ws.cell(row=total_row, column=col_balance, value=f"=SUM({L}2:{L}{last_row})").font = Font(bold=True)
            ws.cell(row=total_row, column=col_balance).number_format = "#,##0"

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    def _style_summary_sheet(ws):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=1, column=c).font = Font(bold=True)

        second_header_row = None
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == "계좌주" and ws.cell(row=r - 1, column=1).value is None:
                second_header_row = r
                break

        if second_header_row:
            for c in range(1, ws.max_column + 1):
                ws.cell(row=second_header_row, column=c).font = Font(bold=True)
            headers2 = {ws.cell(row=second_header_row, column=c).value: c for c in range(1, ws.max_column + 1)}
            for hdr in ["수신합계", "송금합계", "순액(수신-송금)"]:
                if hdr in headers2:
                    col = headers2[hdr]
                    rr = second_header_row + 1
                    while rr <= ws.max_row and ws.cell(row=rr, column=1).value is not None:
                        ws.cell(row=rr, column=col).number_format = "#,##0"
                        rr += 1

    def build_excel_bytes(inter_marked: pd.DataFrame) -> bytes:
        # 1) 1차로 엑셀 작성 (각 시트 + 통합 요약)
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            # 계좌주별 시트
            for owner, sub in inter_marked.groupby("계좌주"):
                sub = sub.sort_values(["상대계좌주", "일시"])
                cols = ["일시", "계좌주", "계좌번호", "상대계좌주", "방향", "거래대금", "잔액", "거래방법", "비고", "상계후보"]
                sub[cols].to_excel(writer, index=False, sheet_name=str(owner))

            # 통합 요약 시트 작성용 데이터
            m = inter_marked[inter_marked["상계후보"]].copy()
            pivot_sum = m.pivot_table(
                index=["계좌주", "상대계좌주"],
                columns="방향",
                values="거래대금",
                aggfunc="sum",
                fill_value=0.0
            ).reset_index()
            if "in" not in pivot_sum.columns:
                pivot_sum["in"] = 0.0
            if "out" not in pivot_sum.columns:
                pivot_sum["out"] = 0.0
            pivot_sum["순액(수신-송금)"] = pivot_sum["in"] - pivot_sum["out"]

            pivot_cnt = m.groupby(["계좌주", "상대계좌주", "방향"])["거래대금"].count().reset_index(name="건수")
            pivot_cnt = pivot_cnt.pivot_table(index=["계좌주", "상대계좌주"], columns="방향", values="건수", fill_value=0).reset_index()
            pivot_cnt = pivot_cnt.rename(columns={"in": "수신건수", "out": "송건수"})

            merged_pairs = pd.merge(pivot_sum, pivot_cnt, on=["계좌주", "상대계좌주"], how="left")
            merged_pairs = merged_pairs.rename(columns={"in": "수신합계", "out": "송금합계"})

            owner_sum = m.pivot_table(
                index=["계좌주"],
                columns="방향",
                values="거래대금",
                aggfunc="sum",
                fill_value=0.0
            ).reset_index()
            if "in" not in owner_sum.columns:
                owner_sum["in"] = 0.0
            if "out" not in owner_sum.columns:
                owner_sum["out"] = 0.0
            owner_sum["순액(수신-송금)"] = owner_sum["in"] - owner_sum["out"]
            owner_sum = owner_sum.rename(columns={"in": "수신합계", "out": "송금합계"})

            merged_pairs.to_excel(writer, index=False, sheet_name="통합", startrow=0)
            owner_sum.to_excel(writer, index=False, sheet_name="통합", startrow=len(merged_pairs) + 3)

        # 2) 스타일 적용 (openpyxl로 재오픈)
        buffer.seek(0)
        wb = load_workbook(buffer)
        for sheet_name in wb.sheetnames:
            if sheet_name == "통합":
                continue
            _style_owner_sheet(wb[sheet_name])
        _style_summary_sheet(wb["통합"])

        # 3) 다시 BytesIO로 저장
        out_buf = io.BytesIO()
        wb.save(out_buf)
        out_buf.seek(0)
        return out_buf.getvalue()
else:
    # openpyxl이 없으면 스타일 없이 xlsxwriter로 저장만 수행
    def build_excel_bytes(inter_marked: pd.DataFrame) -> bytes:
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="xlsxwriter") as writer:
            for owner, sub in inter_marked.groupby("계좌주"):
                sub = sub.sort_values(["상대계좌주", "일시"])
                cols = ["일시", "계좌주", "계좌번호", "상대계좌주", "방향", "거래대금", "잔액", "거래방법", "비고", "상계후보"]
                sub[cols].to_excel(writer, index=False, sheet_name=str(owner))

            m = inter_marked[inter_marked["상계후보"]].copy()
            pivot_sum = m.pivot_table(
                index=["계좌주", "상대계좌주"],
                columns="방향",
                values="거래대금",
                aggfunc="sum",
                fill_value=0.0
            ).reset_index()
            if "in" not in pivot_sum.columns:
                pivot_sum["in"] = 0.0
            if "out" not in pivot_sum.columns:
                pivot_sum["out"] = 0.0
            pivot_sum["순액(수신-송금)"] = pivot_sum["in"] - pivot_sum["out"]

            pivot_cnt = m.groupby(["계좌주", "상대계좌주", "방향"])["거래대금"].count().reset_index(name="건수")
            pivot_cnt = pivot_cnt.pivot_table(index=["계좌주", "상대계좌주"], columns="방향", values="건수", fill_value=0).reset_index()
            pivot_cnt = pivot_cnt.rename(columns={"in": "수신건수", "out": "송건수"})

            merged_pairs = pd.merge(pivot_sum, pivot_cnt, on=["계좌주", "상대계좌주"], how="left")
            merged_pairs = merged_pairs.rename(columns={"in": "수신합계", "out": "송금합계"})

            owner_sum = m.pivot_table(index=["계좌주"], columns="방향", values="거래대금", aggfunc="sum", fill_value=0.0).reset_index()
            if "in" not in owner_sum.columns:
                owner_sum["in"] = 0.0
            if "out" not in owner_sum.columns:
                owner_sum["out"] = 0.0
            owner_sum["순액(수신-송금)"] = owner_sum["in"] - owner_sum["out"]
            owner_sum = owner_sum.rename(columns={"in": "수신합계", "out": "송금합계"})

            merged_pairs.to_excel(writer, index=False, sheet_name="통합", startrow=0)
            owner_sum.to_excel(writer, index=False, sheet_name="통합", startrow=len(merged_pairs) + 3)

        buffer.seek(0)
        return buffer.read()


# --------------------------- Streamlit UI ---------------------------
st.set_page_config(page_title="계좌주간 거래 분석", page_icon="💰", layout="wide")
st.title("💰 계좌주간 거래 분석 도구 (Streamlit)")

with st.sidebar:
    st.subheader("설정")
    amount_tolerance = st.number_input("금액 허용 오차 (원)", value=DEFAULT_AMOUNT_TOLERANCE, step=10_000, min_value=0)
    time_unit = st.selectbox("시간 단위", ["day", "month", "year"], index=["day","month","year"].index(DEFAULT_TIME_UNIT))
    window = st.number_input("시간창 크기", value=DEFAULT_WINDOW, step=1, min_value=1)
    st.caption("'급여/월급/봉급/급여입금/salary'가 들어간 비고는 자동 제외됩니다.")

# 기본 데이터 안내 메시지
st.info("📊 **기본 샘플 데이터가 적용되어 있습니다.** 바로 '분석 실행' 버튼을 눌러보세요! 다른 파일을 업로드하면 해당 파일로 분석됩니다.")

st.markdown("""
**사용 방법**
1) 기본 샘플 데이터로 바로 분석을 시작하거나, 아래에서 직접 Excel 파일을 업로드합니다.  
2) 좌측 사이드바에서 파라미터를 설정합니다.  
3) **분석 실행** 버튼을 누르면 결과를 생성합니다.  
4) **다운로드** 버튼으로 결과 xlsx를 저장하세요.
""")

uploaded = st.file_uploader("거래내역 Excel 파일 업로드 (xlsx 권장)", type=["xlsx", "xls", "xlsm"], accept_multiple_files=False)

# 기본 데이터 로드
if uploaded is None:
    default_file = load_default_data()
    if default_file:
        st.success("✅ 기본 샘플 데이터(계좌내역.xlsx)가 로드되었습니다.")
        current_file = default_file
        current_filename = "계좌내역.xlsx"
    else:
        current_file = None
        current_filename = None
else:
    st.success(f"업로드됨: {uploaded.name}")
    current_file = uploaded
    current_filename = uploaded.name

if current_file is not None:
    with st.expander("미리보기 (상위 10행)", expanded=False):
        try:
            preview_df = _read_excel_any(current_file, current_filename).head(10)
            st.dataframe(preview_df, use_container_width=True)
            # BytesIO 객체의 경우 seek(0)로 위치 초기화
            if hasattr(current_file, 'seek'):
                current_file.seek(0)
        except Exception as e:
            st.warning(str(e))

run = st.button("🔍 분석 실행", type="primary", use_container_width=True)

if run:
    if current_file is None:
        st.error("파일을 불러올 수 없습니다. 인터넷 연결을 확인하거나 직접 파일을 업로드해 주세요.")
    else:
        with st.spinner("분석 중입니다..."):
            try:
                # 1) 데이터 읽기 & 전처리
                df = read_data_from_upload(current_file, current_filename)
                inter = extract_inter_owner_transactions(df)
                inter_marked = match_out_to_in_and_backfill(
                    inter, amount_tolerance=amount_tolerance, time_unit=time_unit, window=window
                )

                # 2) 통합 요약 일부를 화면에서도 보여주기 (엑셀 "통합" 시트의 상세 테이블 형태)
                m = inter_marked[inter_marked["상계후보"]].copy()
                if m.empty:
                    st.info("상계 후보로 매칭된 거래가 없습니다. 파라미터를 조정해 보세요.")
                else:
                    st.subheader("요약 미리보기")
                    
                    # 계좌주-상대계좌주 쌍별 집계 (엑셀의 통합 시트와 동일한 형식)
                    pivot_sum = m.pivot_table(
                        index=["계좌주", "상대계좌주"],
                        columns="방향",
                        values="거래대금",
                        aggfunc="sum",
                        fill_value=0.0
                    ).reset_index()
                    if "in" not in pivot_sum.columns:
                        pivot_sum["in"] = 0.0
                    if "out" not in pivot_sum.columns:
                        pivot_sum["out"] = 0.0
                    pivot_sum["순액(수신-송금)"] = pivot_sum["in"] - pivot_sum["out"]
                    
                    pivot_cnt = m.groupby(["계좌주", "상대계좌주", "방향"])["거래대금"].count().reset_index(name="건수")
                    pivot_cnt = pivot_cnt.pivot_table(
                        index=["계좌주", "상대계좌주"], 
                        columns="방향", 
                        values="건수", 
                        fill_value=0
                    ).reset_index()
                    if "in" in pivot_cnt.columns:
                        pivot_cnt = pivot_cnt.rename(columns={"in": "수신건수"})
                    else:
                        pivot_cnt["수신건수"] = 0
                    if "out" in pivot_cnt.columns:
                        pivot_cnt = pivot_cnt.rename(columns={"out": "송건수"})
                    else:
                        pivot_cnt["송건수"] = 0
                    
                    merged_pairs = pd.merge(pivot_sum, pivot_cnt, on=["계좌주", "상대계좌주"], how="left")
                    merged_pairs = merged_pairs.rename(columns={"in": "수신합계", "out": "송금합계"})
                    
                    # 컬럼 순서 정리
                    display_cols = ["계좌주", "상대계좌주", "수신합계", "송금합계", "순액(수신-송금)", "수신건수", "송건수"]
                    merged_pairs = merged_pairs[[c for c in display_cols if c in merged_pairs.columns]]
                    
                    st.dataframe(merged_pairs, use_container_width=True)

                # 3) 엑셀 Bytes 생성 & 다운로드 버튼
                xlsx_bytes = build_excel_bytes(inter_marked)
                label = "📥 결과 파일 다운로드 (계좌주간_거래_분석.xlsx)"
                if not OPENPYXL_AVAILABLE:
                    st.info("openpyxl 미설치로 인해 셀 스타일 적용 없이 저장됩니다. 'pip install openpyxl' 설치 시 색상/합계 서식이 적용됩니다.")
                st.download_button(
                    label=label,
                    data=xlsx_bytes,
                    file_name="계좌주간_거래_분석.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

                st.success("분석이 완료되었습니다.")

            except Exception as e:
                st.error(f"오류 발생: {e}")