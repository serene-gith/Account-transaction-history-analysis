# -*- coding: utf-8 -*-
import io
import re
import difflib
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import requests
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="ê³„ì¢Œì£¼ê°„ ê±°ë˜ ë¶„ì„(Excel+GitHub)", page_icon="ğŸ§®", layout="wide")
st.title("ğŸ§® ê³„ì¢Œì£¼ê°„ ê±°ë˜ ë¶„ì„ ë„êµ¬")

with st.sidebar:
    st.header("ì„¤ì •")
    st.caption("ìƒê³„(ì™•ë³µ) íŒì • ë° ìˆœìœ ì¶œ ì˜ì‹¬ ê¸°ì¤€ì„ ì¡°ì •í•  ìˆ˜ ìˆì–´ìš”.")
    WINDOW_DAYS = st.number_input("ê¸°ê°„ í—ˆìš©(ì¼)", min_value=0, max_value=60, value=7, step=1)
    AMOUNT_TOL_RATIO = st.number_input("ê¸ˆì•¡ í—ˆìš©ë¹„ìœ¨(%)", min_value=0.0, max_value=10.0, value=2.0, step=0.1) / 100.0
    AMOUNT_TOL_ABS = st.number_input("ê¸ˆì•¡ ì ˆëŒ€í—ˆìš©(ì›)", min_value=0, max_value=10_000_000, value=10_000, step=1_000)
    GIFT_NET_THRESHOLD = st.number_input("ì¦ì—¬ì˜ì‹¬(ìˆœìœ ì¶œ ê¸°ì¤€, ì›)", min_value=0, max_value=10_000_000_000, value=10_000_000, step=1_000_000)
    st.markdown("---")
    st.subheader("GitHubì—ì„œ ê°€ì ¸ì˜¤ê¸°")
    repo_url = st.text_input(
        "ì €ì¥ì†Œ URL",
        value="https://github.com/serene-gith/Account-transaction-history-analysis",
        placeholder="https://github.com/user/repo"
    )
    branch = st.text_input("ë¸Œëœì¹˜", value="main")
    folder_path = st.text_input("í´ë” ê²½ë¡œ(ë£¨íŠ¸ë©´ ë¹„ì›€)", value="")
    gh_token = st.text_input("í† í°(ë¹„ê³µê°œ ì €ì¥ì†Œ)", type="password")
    st.caption("ê³µê°œ ì €ì¥ì†ŒëŠ” í† í° ì—†ì´ ë™ì‘í•©ë‹ˆë‹¤.")

st.markdown("#### 1) íŒŒì¼ ì—…ë¡œë“œ ë˜ëŠ” 2) GitHubì—ì„œ ì„ íƒ")

tab1, tab2 = st.tabs(["ğŸ“¤ ë¡œì»¬ íŒŒì¼ ì—…ë¡œë“œ", "ğŸ™ GitHubì—ì„œ ì„ íƒ"])

# session
for k in ["uploaded_bytes","uploaded_name","gh_ctx"]:
    if k not in st.session_state:
        st.session_state[k] = None

uploaded_bytes = None
uploaded_name = None

with tab1:
    up = st.file_uploader("ê±°ë˜ë‚´ì—­ Excel íŒŒì¼ ì—…ë¡œë“œ (.xlsx/.xlsm/.xls)", type=["xlsx","xlsm","xls"])
    if up is not None:
        uploaded_bytes = up.read()
        uploaded_name = up.name
        st.session_state["uploaded_bytes"] = uploaded_bytes
        st.session_state["uploaded_name"] = uploaded_name
        st.success(f"ì—…ë¡œë“œë¨: {uploaded_name}")
        try:
            preview = pd.read_excel(io.BytesIO(uploaded_bytes), nrows=10, header=None)
            st.expander("ë¯¸ë¦¬ë³´ê¸°(ìƒìœ„ 10í–‰, í—¤ë” íƒì§€ ì „)").dataframe(preview)
        except Exception as e:
            st.warning(f"ë¯¸ë¦¬ë³´ê¸° ì‹¤íŒ¨: {e}")

@st.cache_data(show_spinner=False)
def list_repo_xlsx(repo_url: str, branch: str, folder_path: str, token: Optional[str]):
    parts = repo_url.strip("/").split("/")
    if len(parts) < 2:
        raise ValueError("ì €ì¥ì†Œ URL í˜•ì‹ì´ ì˜¬ë°”ë¥´ì§€ ì•ŠìŠµë‹ˆë‹¤. ì˜ˆ: https://github.com/user/repo")
    owner, repo = parts[-2], parts[-1]
    path = folder_path.strip("/")
    api_url = f"https://api.github.com/repos/{owner}/{repo}/contents/{path}" if path else f"https://api.github.com/repos/{owner}/{repo}/contents"
    params = {"ref": branch}
    headers = {"Accept": "application/vnd.github+json"}
    if token:
        headers["Authorization"] = f"token {token}"
    r = requests.get(api_url, params=params, headers=headers, timeout=30)
    r.raise_for_status()
    items = r.json()
    if not isinstance(items, list):
        items = [items]
    excel_items = []
    for it in items:
        if it.get("type") == "file":
            name = it.get("name", "")
            if name.lower().endswith((".xlsx",".xlsm",".xls")):
                excel_items.append({"name": name, "path": it.get("path")})
    excel_items.sort(key=lambda x: x["name"].lower())
    return owner, repo, excel_items

@st.cache_data(show_spinner=False)
def fetch_file_bytes(owner: str, repo: str, branch: str, file_path: str, token: Optional[str]):
    raw_url = f"https://raw.githubusercontent.com/{owner}/{repo}/{branch}/{file_path}"
    headers = {}
    if token:
        headers["Authorization"] = f"token {token}"
    r = requests.get(raw_url, headers=headers, timeout=60)
    if r.status_code == 404:
        api_url = f"https://api.github.com/repos/{owner}/{repo}/contents/{file_path}"
        params = {"ref": branch}
        api_headers = {"Accept": "application/vnd.github+json"}
        if token:
            api_headers["Authorization"] = f"token {token}"
        res = requests.get(api_url, params=params, headers=api_headers, timeout=60)
        res.raise_for_status()
        data = res.json()
        download_url = data.get("download_url")
        if not download_url:
            raise RuntimeError("download_urlì„ ê°€ì ¸ì˜¤ì§€ ëª»í–ˆìŠµë‹ˆë‹¤.")
        r = requests.get(download_url, timeout=60)
    r.raise_for_status()
    return r.content

with tab2:
    if st.button("ğŸ” ëª©ë¡ ë¶ˆëŸ¬ì˜¤ê¸°", use_container_width=True) and repo_url.strip():
        try:
            owner, repo, excel_list = list_repo_xlsx(repo_url, branch, folder_path, gh_token or None)
            st.session_state["gh_ctx"] = (owner, repo, excel_list)
            st.success("ëª©ë¡ì„ ë¶ˆëŸ¬ì™”ìŠµë‹ˆë‹¤.")
        except Exception as e:
            st.error(f"GitHub ëª©ë¡ ì˜¤ë¥˜: {e}")

    if st.session_state.get("gh_ctx"):
        owner, repo, excel_list = st.session_state["gh_ctx"]
        if not excel_list:
            st.warning("ì—‘ì…€ íŒŒì¼ì´ ì—†ìŠµë‹ˆë‹¤. ë¸Œëœì¹˜/í´ë” ê²½ë¡œë¥¼ í™•ì¸í•˜ì„¸ìš”.")
        else:
            names = [x["name"] for x in excel_list]
            pick = st.selectbox("íŒŒì¼ ì„ íƒ", names, index=0)
            if st.button("â¬‡ï¸ ì´ íŒŒì¼ ê°€ì ¸ì˜¤ê¸°", type="primary"):
                try:
                    item = next(x for x in excel_list if x["name"] == pick)
                    data = fetch_file_bytes(owner, repo, branch, item["path"], gh_token or None)
                    uploaded_bytes = data
                    uploaded_name = pick
                    st.session_state["uploaded_bytes"] = data
                    st.session_state["uploaded_name"] = pick
                    st.success(f"ê°€ì ¸ì˜´: {pick}")
                    prev = pd.read_excel(io.BytesIO(uploaded_bytes), nrows=10, header=None)
                    st.expander("ë¯¸ë¦¬ë³´ê¸°(ìƒìœ„ 10í–‰, í—¤ë” íƒì§€ ì „)").dataframe(prev)
                except Exception as e:
                    st.error(f"ê°€ì ¸ì˜¤ê¸° ì˜¤ë¥˜: {e}")

# ----------- Analysis core -----------
SCAN_HEADER_ROWS = 20

SYN = {
    # ë‚ ì§œ/ì‹œê°„/ì„¤ëª…/ê¸ˆì•¡/ì†Œìœ ì/ìƒëŒ€/ê³„ì¢Œ
    "date": ["ê±°ë˜ì¼","ê±°ë˜ì¼ì","ì¼ì","date","transdate","ë‚ ì§œ"],
    "time": ["ê±°ë˜ì‹œê°","ì‹œê°„","time","ì‹œê°"],
    "owner": ["ê³„ì¢Œì£¼","ì˜ˆê¸ˆì£¼","ì†Œìœ ì","owner","accountowner"],
    "owner_acct": ["ê³„ì¢Œë²ˆí˜¸","ê³„ì¢Œ","account","owneraccount","fromaccount","ì¶œê¸ˆê³„ì¢Œ","ì¶œê¸ˆê³„ì¢Œë²ˆí˜¸"],
    "cp_holder": ["ìƒëŒ€ê³„ì¢Œì£¼","ìƒëŒ€ë°©","ê±°ë˜ìƒëŒ€","ìˆ˜ì·¨ì¸","ë°›ëŠ”ì´","ë°›ëŠ”ì‚¬ëŒ","ì…ê¸ˆê³„ì¢Œì£¼","counterparty","ê±°ë˜ìƒëŒ€ë°©"],
    "cp_acct": ["ìƒëŒ€ê³„ì¢Œë²ˆí˜¸","ìƒëŒ€ê³„ì¢Œ","counterpartyaccount","ì…ê¸ˆê³„ì¢Œ","ì…ê¸ˆê³„ì¢Œë²ˆí˜¸","ë°›ëŠ”ê³„ì¢Œ","ë°›ëŠ”ê³„ì¢Œë²ˆí˜¸"],
    "desc": ["ê±°ë˜ë‚´ì—­","ì ìš”","ë©”ëª¨","ë‚´ìš©","ë¹„ê³ ","summary","description","ë‚´ì—­","ë¹„ê³ ê±°ë˜ë‚´ìš©","ë¹„ê³ ì¶”ê°€"],
    "amount": ["ê¸ˆì•¡","ê±°ë˜ê¸ˆì•¡","ì´ì²´ê¸ˆì•¡","amount","amt","ê¸ˆì•¡(ì›)","ê±°ë˜ëŒ€ê¸ˆ","ê±°ë˜ëŒ€ê¸ˆì›","ì…ê¸ˆê¸ˆì•¡","ì¶œê¸ˆê¸ˆì•¡","ì…ê¸ˆì›","ì¶œê¸ˆì›"],
    "credit": ["ì…ê¸ˆ","ì…ê¸ˆì•¡","ìˆ˜ì·¨ê¸ˆì•¡","ë°›ì€ê¸ˆì•¡","credit","cr"],
    "debit": ["ì¶œê¸ˆ","ì¶œê¸ˆì•¡","ì†¡ê¸ˆì•¡","ë³´ë‚¸ê¸ˆì•¡","debit","dr"],
    "method": ["ê±°ë˜ë°©ë²•","êµ¬ë¶„","ì…ì¶œê¸ˆêµ¬ë¶„","ê±°ë˜êµ¬ë¶„"],
    "from_holder": ["ì¶œê¸ˆê³„ì¢Œì£¼","ë³´ë‚¸ì‚¬ëŒ","ì†¡ê¸ˆì¸","fromholder","ì¶œê¸ˆê³„ì¢Œëª…"],
    "from_acct": ["ì¶œê¸ˆê³„ì¢Œ","ì¶œê¸ˆê³„ì¢Œë²ˆí˜¸","fromaccount","ë³´ë‚¸ê³„ì¢Œ"],
    "to_holder": ["ì…ê¸ˆê³„ì¢Œì£¼","ë°›ëŠ”ì‚¬ëŒ","ìˆ˜ì·¨ì¸ëª…","toholder","ì…ê¸ˆê³„ì¢Œëª…"],
    "to_acct": ["ì…ê¸ˆê³„ì¢Œ","ì…ê¸ˆê³„ì¢Œë²ˆí˜¸","toaccount","ë°›ëŠ”ê³„ì¢Œ"],
}

def _clean(s: str) -> str:
    s = str(s)
    s = s.replace("\n","").replace("\r","")
    s = re.sub(r"\s+", "", s)
    s = s.replace("(", "").replace(")", "")
    return s.lower()

def _to_number(x):
    if pd.isna(x):
        return np.nan
    s = str(x).strip()
    if re.match(r"^\(.*\)$", s):
        s = "-" + s[1:-1]
    s = s.replace(",", "")
    try:
        return float(s)
    except Exception:
        return pd.to_numeric(s, errors="coerce")

def _best_match(col: str, candidates: List[str]) -> Optional[str]:
    if col in candidates:
        return col
    for c in candidates:
        if c in col or col in c:
            return c
    m = difflib.get_close_matches(col, candidates, n=1, cutoff=0.72)
    return m[0] if m else None

def _find_header(df: pd.DataFrame) -> pd.DataFrame:
    if not df.empty and all(isinstance(c, str) for c in df.columns):
        return df
    best_row, best_score = None, -1
    for r in range(min(SCAN_HEADER_ROWS, len(df))):
        row_vals = df.iloc[r].astype(str).tolist()
        cleaned = [_clean(v) for v in row_vals]
        tokens = sum(1 for v in cleaned if _best_match(v, sum(SYN.values(), [])) is not None)
        if tokens > best_score:
            best_score, best_row = tokens, r
    if best_row is None:
        return df
    new_cols = [_clean(x) for x in df.iloc[best_row].tolist()]
    df2 = df.iloc[best_row+1:].copy()
    df2.columns = new_cols
    df2.reset_index(drop=True, inplace=True)
    return df2

def _map_columns(cols: List[str]) -> Dict[str, Optional[str]]:
    def pick(keys: List[str]) -> Optional[str]:
        for key in keys:
            cands = [_clean(x) for x in SYN.get(key, [])]
            for col in cols:
                bm = _best_match(col, cands)
                if bm:
                    return col
        return None
    return {
        "owner": pick(["owner"]),
        "owner_acct": pick(["owner_acct"]),
        "cp_holder": pick(["cp_holder"]),
        "cp_acct": pick(["cp_acct"]),
        "amount": pick(["amount"]),
        "desc": pick(["desc"]),
        "date": pick(["date"]),
        "time": pick(["time"]),
        "method": pick(["method"]),
        "from_holder": pick(["from_holder"]),
        "from_acct": pick(["from_acct"]),
        "to_holder": pick(["to_holder"]),
        "to_acct": pick(["to_acct"]),
        "credit": pick(["credit"]),
        "debit": pick(["debit"]),
    }

def _infer_direction(val: str) -> Optional[str]:
    if not isinstance(val, str):
        return None
    v = val.strip()
    if "ì…" in v:  # ì…ê¸ˆ/ì…ê¸ˆì•¡ ë“±
        return "credit"
    if "ì¶œ" in v:  # ì¶œê¸ˆ/ì¶œê¸ˆì•¡ ë“±
        return "debit"
    return None

# descì—ì„œ ìƒëŒ€ì •ë³´ ì¶”ì¶œ ì‹œë„
CP_PATTERNS = [
    r"(?:ìƒëŒ€|ê±°ë˜ìƒëŒ€|ìƒëŒ€ê³„ì¢Œì£¼|ì…ê¸ˆê³„ì¢Œì£¼|ë°›ëŠ”ì‚¬ëŒ|ìˆ˜ì·¨ì¸|ë°›ëŠ”ë¶„)[:ï¼š\s]\s*([^\s,/;|]+)",
    r"(?:ë³´ë‚¸ì‚¬ëŒ|ì†¡ê¸ˆì¸|ì¶œê¸ˆê³„ì¢Œì£¼)[:ï¼š\s]\s*([^\s,/;|]+)",
    r"(?:ìƒëŒ€ê³„ì¢Œë²ˆí˜¸|ì…ê¸ˆê³„ì¢Œ|ë°›ëŠ”ê³„ì¢Œ)[:ï¼š\s]\s*([\d\-]+)",
]

def _extract_counterparty(desc: str) -> Dict[str, Optional[str]]:
    if not isinstance(desc, str):
        return {"cp_holder": None, "cp_acct": None}
    holder = None
    acct = None
    for pat in CP_PATTERNS:
        m = re.search(pat, desc)
        if m:
            if "ê³„ì¢Œ" in pat and m.group(1):
                acct = m.group(1)
            else:
                holder = holder or m.group(1)
    return {"cp_holder": holder, "cp_acct": acct}

def load_and_standardize_from_bytes(b: bytes) -> Tuple[pd.DataFrame, Dict[str,str], List[str]]:
    raw = pd.read_excel(io.BytesIO(b), header=None)
    df = _find_header(raw.copy())
    cols = [_clean(c) for c in df.columns]
    mapping = _map_columns(cols)

    # combine date + time
    if mapping["date"] and mapping["time"] and mapping["time"] in df.columns:
        dt = pd.to_datetime(df[mapping["date"]], errors="coerce")
        tm = pd.to_datetime(df[mapping["time"]].astype(str), format="%H:%M:%S", errors="coerce").dt.time
        date_series = pd.to_datetime(dt.dt.date.astype(str) + " " + df[mapping["time"]].astype(str), errors="coerce")
        date_col = date_series
    else:
        date_col = pd.to_datetime(df[mapping["date"]], errors="coerce") if mapping["date"] else pd.NaT

    has_single_amount = mapping["amount"] in cols if mapping["amount"] else False
    has_split_amounts = (mapping["credit"] in cols if mapping["credit"] else False) or (mapping["debit"] in cols if mapping["debit"] else False)

    # CASE 1: ëª…ì‹œì  from/toê°€ ìˆëŠ” Bí˜•
    if all(mapping.get(k) for k in ["from_holder","from_acct","to_holder","to_acct"]) and (has_single_amount or has_split_amounts):
        df_std = pd.DataFrame({
            "date": date_col,
            "from_holder": df[mapping["from_holder"]].astype(str).str.strip(),
            "from_acct": df[mapping["from_acct"]].astype(str).str.strip(),
            "to_holder": df[mapping["to_holder"]].astype(str).str.strip(),
            "to_acct": df[mapping["to_acct"]].astype(str).str.strip(),
            "desc": df[mapping["desc"]].astype(str) if mapping["desc"] else "",
        })
        if has_single_amount:
            df_std["amount"] = df[mapping["amount"]].apply(_to_number)
            amount_info = f"amount â† {mapping['amount']}"
        else:
            cr = df[mapping["credit"]].apply(_to_number) if mapping["credit"] else 0.0
            dr = df[mapping["debit"]].apply(_to_number) if mapping["debit"] else 0.0
            df_std["amount"] = cr.where(cr.notna() & (cr != 0), 0) - dr.where(dr.notna() & (dr != 0), 0)
            amount_info = f"amount â† +{mapping.get('credit')} -{mapping.get('debit')}"
        schema = "B(from/to)"

    # CASE 2: owner/cpí˜•(A) ë˜ëŠ” ownerë§Œ ìˆê³  desc/ë°©ë²•ìœ¼ë¡œ ë³´ê°•
    elif mapping["owner"] and mapping["owner_acct"] and (has_single_amount or has_split_amounts):
        owner = df[mapping["owner"]].astype(str).str.strip()
        owner_acct = df[mapping["owner_acct"]].astype(str).str.strip()
        desc_col = df[mapping["desc"]].astype(str) if mapping["desc"] else ""
        # counterparty ìš°ì„  ì‚¬ìš©, ì—†ìœ¼ë©´ descì—ì„œ ì¶”ì¶œ
        cp_holder = df[mapping["cp_holder"]].astype(str).str.strip() if mapping["cp_holder"] else None
        cp_acct = df[mapping["cp_acct"]].astype(str).str.strip() if mapping["cp_acct"] else None
        if cp_holder is None or (isinstance(cp_holder, pd.Series) and cp_holder.isna().all()):
            ext = desc_col.apply(_extract_counterparty)
            cp_holder = ext.apply(lambda d: d["cp_holder"])
            cp_acct = ext.apply(lambda d: d["cp_acct"])

        if has_single_amount:
            amt = df[mapping["amount"]].apply(_to_number)
            mdir = df[mapping["method"]] if mapping["method"] else None
            if mdir is not None:
                dir_series = mdir.astype(str).apply(_infer_direction)
                # dirì— ë”°ë¼ ë¶€í˜¸ ì ìš©: ì…ê¸ˆ=+, ì¶œê¸ˆ=- (owner ê¸°ì¤€)
                sign = dir_series.map({"credit": 1, "debit": -1}).fillna(0)
                amt_signed = amt.copy()
                amt_signed[sign == -1] = -abs(amt_signed[sign == -1])
                amt_signed[sign == 1] = abs(amt_signed[sign == 1])
                amt = amt_signed
            amount_info = f"amount â† {mapping['amount']} (method used: {mapping.get('method')})"
        else:
            cr = df[mapping["credit"]].apply(_to_number) if mapping["credit"] else 0.0
            dr = df[mapping["debit"]].apply(_to_number) if mapping["debit"] else 0.0
            amt = cr.where(cr.notna() & (cr != 0), 0) - dr.where(dr.notna() & (dr != 0), 0)
            amount_info = f"amount â† +{mapping.get('credit')} -{mapping.get('debit')}"

        # ë°©í–¥ ê²°ì •: ê¸ˆì•¡ ë¶€í˜¸ ë˜ëŠ” method, ì—†ìœ¼ë©´ owner->cp ê°€ì •
        if isinstance(amt, pd.Series):
            to_is_owner = amt >= 0  # ì–‘ìˆ˜ë©´ ownerê°€ ë°›ìŒ â‡’ cp -> owner
            from_holder = np.where(to_is_owner, cp_holder, owner)
            to_holder = np.where(to_is_owner, owner, cp_holder)
            from_acct = np.where(to_is_owner, cp_acct, owner_acct)
            to_acct = np.where(to_is_owner, owner_acct, cp_acct)
        else:
            from_holder, to_holder = owner, cp_holder
            from_acct, to_acct = owner_acct, cp_acct

        df_std = pd.DataFrame({
            "date": date_col,
            "from_holder": from_holder,
            "from_acct": from_acct,
            "to_holder": to_holder,
            "to_acct": to_acct,
            "amount": amt,
            "desc": desc_col,
        })
        schema = "A(owner/desc-augmented)"

    elif all(k in df.columns for k in ["date","from_holder","from_acct","to_holder","to_acct","amount"]):
        df_std = pd.DataFrame({
            "date": pd.to_datetime(df["date"], errors="coerce"),
            "from_holder": df["from_holder"].astype(str).str.strip(),
            "from_acct": df["from_acct"].astype(str).str.strip(),
            "to_holder": df["to_holder"].astype(str).str.strip(),
            "to_acct": df["to_acct"].astype(str).str.strip(),
            "amount": pd.to_numeric(df["amount"], errors="coerce"),
            "desc": df["desc"].astype(str) if "desc" in df.columns else "",
        })
        amount_info = "already standardized"
        schema = "C(standardized)"
    else:
        raise ValueError(f"ì»¬ëŸ¼ ìë™ ì¸ì‹ ì‹¤íŒ¨. ê°ì§€ëœ ì»¬ëŸ¼: {cols}")

    df_std = df_std.dropna(subset=["from_holder","to_holder","amount"]).copy()
    if df_std["date"].notna().any():
        df_std.sort_values(["date"], inplace=True, kind="mergesort")
    df_std.reset_index(drop=True, inplace=True)

    mapping_summary = {
        "schema_detected": schema,
        "amount_info": amount_info if 'amount_info' in locals() else "",
        **mapping
    }
    return df_std, mapping_summary, cols

# --- matching & output (same as previous version) ---
@dataclass
class Match:
    i: int
    j: int
    reason: str

def _amount_close(a: float, b: float, tol_abs: float, tol_ratio: float) -> bool:
    if pd.isna(a) or pd.isna(b):
        return False
    if abs(a - b) <= tol_abs:
        return True
    base = max(abs(a), abs(b), 1.0)
    return abs(a - b) / base <= tol_ratio

def find_matches(df: pd.DataFrame) -> List[Match]:
    matches: List[Match] = []
    used = set()
    df2 = df.copy()
    df2["pair_key"] = df2.apply(lambda r: frozenset([r["from_holder"], r["to_holder"]]), axis=1)
    for key, g in df2.groupby("pair_key", sort=False):
        if len(key) != 2:
            continue
        a, b = list(key)
        pos = g[(g["from_holder"] == a) & (g["to_holder"] == b)]
        neg = g[(g["from_holder"] == b) & (g["to_holder"] == a)]
        for i in pos.index:
            if i in used:
                continue
            for j in neg.index:
                if j in used:
                    continue
                if not _amount_close(df.loc[i,"amount"], df.loc[j,"amount"], AMOUNT_TOL_ABS, AMOUNT_TOL_RATIO):
                    continue
                d1, d2 = df.loc[i,"date"], df.loc[j,"date"]
                if pd.notna(d1) and pd.notna(d2):
                    if abs((d2 - d1).days) > WINDOW_DAYS:
                        continue
                    reason = f"ê¸ˆì•¡Â·ê¸°ê°„ ìœ ì‚¬(â‰¤{WINDOW_DAYS}ì¼)"
                else:
                    reason = "ê¸ˆì•¡ ìœ ì‚¬(ë‚ ì§œ ì—†ìŒ)"
                used.update([i, j])
                matches.append(Match(i=i, j=j, reason=reason))
                break
    return matches

def build_owner_sheets(df: pd.DataFrame, matches: List[Match]):
    match_id = pd.Series(index=df.index, dtype="Int64")
    reason = pd.Series(index=df.index, dtype="string")
    for k, m in enumerate(matches, start=1):
        match_id.loc[m.i] = k
        match_id.loc[m.j] = k
        reason.loc[m.i] = m.reason
        reason.loc[m.j] = m.reason

    df2 = df.copy()
    df2["match_id"] = match_id
    df2["match_reason"] = reason
    df2["counterparty_holder"] = df2["to_holder"]
    df2["counterparty_account"] = df2["to_acct"]

    owners = sorted(set(df2["from_holder"]).union(set(df2["to_holder"])))
    sheets: Dict[str, pd.DataFrame] = {}
    base_cols = ["date","from_holder","from_acct","to_holder","to_acct","counterparty_holder","counterparty_account","amount","desc","match_id","match_reason"]
    for owner in owners:
        sub = df2[(df2["from_holder"] == owner) | (df2["to_holder"] == owner)].copy()
        sub = sub[sub["from_holder"] != sub["to_holder"]]
        if sub["date"].notna().any():
            sub = sub.sort_values(["date","amount"], kind="mergesort")
        else:
            sub = sub.sort_values(["amount"], kind="mergesort")
        present = [c for c in base_cols if c in sub.columns]
        sheets[owner] = sub[present].reset_index(drop=True)
    return sheets, df2

def make_summary(df: pd.DataFrame) -> pd.DataFrame:
    pairs = df.groupby(["from_holder", "to_holder"], dropna=False)["amount"].sum().reset_index()
    pairs = pairs[pairs["from_holder"] != pairs["to_holder"]].copy()
    pairs["pair_key"] = pairs.apply(lambda r: tuple(sorted([r["from_holder"], r["to_holder"]])), axis=1)
    recs = []
    for key, g in pairs.groupby("pair_key"):
        a, b = key
        a2b = g[(g["from_holder"] == a) & (g["to_holder"] == b)]["amount"].sum()
        b2a = g[(g["from_holder"] == b) & (g["to_holder"] == a)]["amount"].sum()
        net = a2b - b2a
        donor = a if net > 0 else b
        receiver = b if net > 0 else a
        suspicious = abs(net) >= GIFT_NET_THRESHOLD
        recs.append({
            "holder_A": a,
            "holder_B": b,
            "A->B_total": a2b,
            "B->A_total": b2a,
            "net_A_to_B": net,
            "possible_donor": donor if suspicious else "",
            "possible_receiver": receiver if suspicious else "",
            "gift_risk_hint": "ìˆœìœ ì¶œ ê³¼ë‹¤" if suspicious else ""
        })
    return pd.DataFrame(recs).sort_values("net_A_to_B", ascending=False)

def render_and_download(df: pd.DataFrame, sheets: Dict[str, pd.DataFrame], summary: pd.DataFrame, mapping_summary: Dict[str,str], detected_cols: List[str]) -> bytes:
    yellow = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        summary.to_excel(writer, index=False, sheet_name="Summary")
        for owner, df_owner in sheets.items():
            sheet_name = owner[:31] or "Unknown"
            df_owner.to_excel(writer, index=False, sheet_name=sheet_name)
        df.to_excel(writer, index=False, sheet_name="_Master")
        map_df = pd.DataFrame([mapping_summary])
        det_df = pd.DataFrame({"detected_columns": detected_cols})
        map_df.to_excel(writer, index=False, sheet_name="Mapping")
        det_df.to_excel(writer, index=False, sheet_name="Mapping", startrow=map_df.shape[0]+2)
    buffer.seek(0)
    wb = load_workbook(buffer)
    for ws in wb.worksheets:
        if ws.title == "Mapping":
            for col in range(1, ws.max_column + 1):
                max_len = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1))
                ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 60)
            continue
        if ws.title == "Summary":
            for col in range(1, ws.max_column + 1):
                max_len = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1))
                ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 40)
            continue
        headers = {cell.value: idx for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), start=1)}
        mid_col = headers.get("match_id")
        if mid_col:
            for row in ws.iter_rows(min_row=2):
                val = row[mid_col - 1].value
                if val not in (None, "", 0):
                    for cell in row:
                        cell.fill = yellow
        for col in range(1, ws.max_column + 1):
            max_len = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1))
            ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 40)
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()

st.markdown("---")
run_col, info_col = st.columns([1,2])
with run_col:
    run = st.button("ğŸ” ë¶„ì„ ì‹¤í–‰", type="primary", use_container_width=True)
with info_col:
    st.info("íŒŒì¼ì„ ì—…ë¡œë“œí•˜ê±°ë‚˜ GitHubì—ì„œ ì„ íƒí•œ ë’¤ **ë¶„ì„ ì‹¤í–‰**ì„ ëˆ„ë¥´ë©´ ê²°ê³¼ XLSXê°€ ìƒì„±ë©ë‹ˆë‹¤.")

if run:
    uploaded_bytes = st.session_state.get("uploaded_bytes") or uploaded_bytes
    uploaded_name = st.session_state.get("uploaded_name") or uploaded_name
    if not uploaded_bytes:
        st.error("ë¶„ì„í•  íŒŒì¼ì´ ì—†ìŠµë‹ˆë‹¤. ì—…ë¡œë“œí•˜ê±°ë‚˜ GitHubì—ì„œ ì„ íƒí•˜ì„¸ìš”.")
        st.stop()
    try:
        df, mapping_summary, detected_cols = load_and_standardize_from_bytes(uploaded_bytes)
        matches = find_matches(df)
        sheets, master = build_owner_sheets(df, matches)
        summary = make_summary(df)
        result_bytes = render_and_download(master, sheets, summary, mapping_summary, detected_cols)
        st.success("ë¶„ì„ ì™„ë£Œ! ì•„ë˜ ë²„íŠ¼ìœ¼ë¡œ ê²°ê³¼ íŒŒì¼ì„ ì €ì¥í•˜ì„¸ìš”.")
        fname = (uploaded_name or "ê²°ê³¼") + "_ë¶„ì„ê²°ê³¼.xlsx"
        st.download_button("ğŸ’¾ ê²°ê³¼ XLSX ë‹¤ìš´ë¡œë“œ", data=result_bytes, file_name=fname, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        st.dataframe(summary, use_container_width=True, height=300)
    except Exception as e:
        st.exception(e)
        st.error("ë¶„ì„ ì¤‘ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤. Mapping ì‹œíŠ¸ë¥¼ ì°¸ê³ í•´ ì»¬ëŸ¼ëª…ì„ í™•ì¸í•´ ì£¼ì„¸ìš”.")
