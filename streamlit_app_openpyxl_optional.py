# -*- coding: utf-8 -*-
"""
Streamlit ë‹¨ì¼ íŒŒì¼ ë²„ì „: ê³„ì¢Œì£¼ê°„ ê±°ë˜ ë¶„ì„ ë„êµ¬
- ì—…ë¡œë“œí•œ Excel(xlsx, xls) íŒŒì¼ì„ ì½ì–´ì„œ ê³„ì¢Œì£¼ ê°„ ìƒê³„ í›„ë³´ë¥¼ ë§¤ì¹­í•˜ê³  ê²°ê³¼ë¥¼ ë‹¤ìš´ë¡œë“œë¡œ ì œê³µí•©ë‹ˆë‹¤.
- openpyxl ë¯¸ì„¤ì¹˜ í™˜ê²½ì—ì„œë„ ì•±ì´ ì£½ì§€ ì•Šë„ë¡ ì•ˆì „í•˜ê²Œ ë™ì‘í•©ë‹ˆë‹¤(ìŠ¤íƒ€ì¼/ì„œì‹ì€ openpyxl ìˆì„ ë•Œë§Œ ì ìš©).

í•„ìˆ˜ íŒ¨í‚¤ì§€ (ì˜ˆì‹œ):
    pip install streamlit pandas

ì„ íƒ íŒ¨í‚¤ì§€ (ìˆìœ¼ë©´ ë” ì¢‹ìŒ):
    pip install openpyxl            # xlsx ì½ê¸°/ì“°ê¸° + ìŠ¤íƒ€ì¼ ì ìš©ì— í•„ìš”
    pip install "xlrd==1.2.0"      # êµ¬í˜• .xls ì½ê¸°ì— í•„ìš”
"""

import io
import re
import pandas as pd
import numpy as np
import streamlit as st

# -------------------- openpyxl(ì„ íƒ) ì§€ì—° ì„í¬íŠ¸ --------------------
try:
    from openpyxl import load_workbook  # type: ignore
    from openpyxl.styles import PatternFill, Font  # type: ignore
    from openpyxl.utils import get_column_letter  # type: ignore
    OPENPYXL_AVAILABLE = True
except Exception:  # ëª¨ë“ˆì´ ì—†ê±°ë‚˜ ë¶ˆëŸ¬ì˜¤ê¸° ì‹¤íŒ¨í•´ë„ ì•±ì€ ê³„ì† ë™ì‘
    OPENPYXL_AVAILABLE = False

# ===================== ê¸°ë³¸ ì„¤ì •ê°’ (UIì—ì„œ ë³€ê²½) =====================
DEFAULT_AMOUNT_TOLERANCE = 100_000   # ê¸ˆì•¡ í—ˆìš© ì˜¤ì°¨ (ì›)
DEFAULT_TIME_UNIT = "year"           # "day" | "month" | "year"
DEFAULT_WINDOW = 10                  # ì‹œê°„ì°½ í¬ê¸° (ì˜ˆ: 3ê°œì›”)
# ====================================================================

IN_METHOD_PATTERN = r"(ì´ì²´|ì˜ˆê¸ˆ|ëŒ€ì²´)"
EXCLUDE_NOTE_PATTERN = r"(ê¸‰ì—¬|ì›”ê¸‰|ë´‰ê¸‰|ê¸‰ì—¬ì…ê¸ˆ|salary)"

# -------------------------- ìœ í‹¸ í•¨ìˆ˜ --------------------------
def _to_num(x):
    try:
        return float(str(x).replace(",", "").strip())
    except Exception:
        return np.nan


def _read_excel_any(uploaded_file, filename: str) -> pd.DataFrame:
    """ì—…ë¡œë“œí•œ íŒŒì¼ì„ í™•ì¥ì/ì—”ì§„ì— ë§ê²Œ DataFrameìœ¼ë¡œ ì½ìŠµë‹ˆë‹¤.
    - .xlsx/.xlsm: openpyxl ê¶Œì¥(ì—†ìœ¼ë©´ ì•ˆë‚´ ë©”ì‹œì§€)
    - .xls: xlrd(1.2.0 í•„ìš”)
    """
    name = (filename or "").lower()
    if name.endswith(".xls") and not name.endswith(".xlsx"):
        # xlrd 2.xëŠ” xls ë¯¸ì§€ì› â†’ 1.2.0 ì„¤ì¹˜ í•„ìš”
        try:
            return pd.read_excel(uploaded_file, dtype=str, engine="xlrd")
        except Exception as e:
            raise RuntimeError(
                "*.xls íŒŒì¼ì„ ì½ìœ¼ë ¤ë©´ 'xlrd==1.2.0' ì„¤ì¹˜ê°€ í•„ìš”í•©ë‹ˆë‹¤.\n"
                "ëª…ë ¹: pip install \"xlrd==1.2.0\"\n"
                f"ì›ì¸: {e}"
            )
    else:
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError(
                "xlsx/xlsm íŒŒì¼ì„ ì²˜ë¦¬í•˜ë ¤ë©´ openpyxlì´ í•„ìš”í•©ë‹ˆë‹¤.\n"
                "ëª…ë ¹: pip install openpyxl"
            )
        # openpyxlì´ ìˆì„ ë•Œì—ë§Œ ì—”ì§„ ì§€ì •
        return pd.read_excel(uploaded_file, dtype=str, engine="openpyxl")


# ---------------------- ì›ë³¸ ë¡œì§ ì´ì‹ ----------------------

def read_data_from_upload(uploaded_file, filename: str) -> pd.DataFrame:
    df = _read_excel_any(uploaded_file, filename).rename(columns={
        "ê³„ì¢Œì£¼":"ê³„ì¢Œì£¼","ê³„ì¢Œë²ˆí˜¸":"ê³„ì¢Œë²ˆí˜¸","ê±°ë˜ì¼ì":"ê±°ë˜ì¼ì","ê±°ë˜ì‹œê°":"ê±°ë˜ì‹œê°",
        "ê±°ë˜ëŒ€ê¸ˆ(ì›)":"ê±°ë˜ëŒ€ê¸ˆ","ì”ì•¡(ì›)":"ì”ì•¡","ê±°ë˜ë°©ë²•":"ê±°ë˜ë°©ë²•","ë¹„ê³ (ê±°ë˜ë‚´ìš©)":"ë¹„ê³ "
    })
    for c in ["ê³„ì¢Œì£¼","ê³„ì¢Œë²ˆí˜¸","ê±°ë˜ë°©ë²•","ë¹„ê³ "]:
        if c in df.columns:
            df[c] = df[c].astype(str).str.strip()
    # ë‚ ì§œ/ì‹œê° íŒŒì‹±
    df["ê±°ë˜ì¼ì"] = pd.to_datetime(df.get("ê±°ë˜ì¼ì"), errors="coerce").dt.date
    t = pd.to_datetime(df.get("ê±°ë˜ì‹œê°"), format="%H:%M:%S", errors="coerce")
    df["ê±°ë˜ì‹œê°"] = t.dt.time
    df["ì¼ì‹œ"] = pd.to_datetime(df["ê±°ë˜ì¼ì"].astype(str) + " " + df["ê±°ë˜ì‹œê°"].astype(str), errors="coerce")
    # ê¸ˆì•¡/ì”ì•¡ ìˆ˜ì¹˜í™”
    df["ê±°ë˜ëŒ€ê¸ˆ"] = df.get("ê±°ë˜ëŒ€ê¸ˆ", np.nan).apply(_to_num)
    df["ì”ì•¡"] = df.get("ì”ì•¡", np.nan).apply(_to_num)
    # í•„ìˆ˜ ì»¬ëŸ¼ ê²°ì¸¡ ì œê±°
    return df.dropna(subset=["ê³„ì¢Œì£¼","ì¼ì‹œ","ê±°ë˜ëŒ€ê¸ˆ"]).reset_index(drop=True)


def parse_counterparty_and_direction(note: str):
    if not isinstance(note, str):
        return None, None
    s = re.sub(r"\s+", "", note)
    m_out = re.search(r"([ê°€-í£]{2,})ì—ê²Œì†¡ê¸ˆ", s)
    if m_out:
        return m_out.group(1), "out"
    return None, None


def extract_inter_owner_transactions(df: pd.DataFrame) -> pd.DataFrame:
    parsed = df.get("ë¹„ê³ ", pd.Series([None] * len(df))).apply(parse_counterparty_and_direction)
    inter = df.copy()
    inter["ìƒëŒ€ê³„ì¢Œì£¼"] = [p[0] for p in parsed]
    inter["ë°©í–¥"] = [p[1] for p in parsed]
    inter = inter[inter.get("ê±°ë˜ë°©ë²•", "").astype(str).str.contains("ì´ì²´|ëŒ€ì²´|ì˜ˆê¸ˆ", na=False)].copy()
    inter = inter[[
        "ê³„ì¢Œì£¼","ê³„ì¢Œë²ˆí˜¸","ì¼ì‹œ","ê±°ë˜ì¼ì","ê±°ë˜ì‹œê°","ê±°ë˜ëŒ€ê¸ˆ","ì”ì•¡","ê±°ë˜ë°©ë²•","ë¹„ê³ ","ìƒëŒ€ê³„ì¢Œì£¼","ë°©í–¥"
    ]].sort_values(["ì¼ì‹œ","ê³„ì¢Œì£¼"]).reset_index(drop=True)
    inter["orig_index"] = inter.index
    inter["ì œì™¸ë¹„ê³ "] = inter["ë¹„ê³ "].fillna("").str.contains(EXCLUDE_NOTE_PATTERN, case=False, regex=True)
    # ì œì™¸ ê±´ì€ ë°©í–¥/ìƒëŒ€ê³„ì¢Œì£¼ ë¹„ìš°ê¸°
    inter.loc[inter["ì œì™¸ë¹„ê³ "], ["ë°©í–¥","ìƒëŒ€ê³„ì¢Œì£¼"]] = None
    return inter


def time_within_window(ts1: pd.Timestamp, ts2: pd.Timestamp, time_unit: str, window: int) -> bool:
    if pd.isna(ts1) or pd.isna(ts2):
        return False
    if time_unit == "day":
        return abs(ts2 - ts1) <= pd.Timedelta(days=window)
    elif time_unit == "month":
        return abs((ts2.year - ts1.year)*12 + (ts2.month - ts1.month)) <= window
    elif time_unit == "year":
        return abs(ts2.year - ts1.year) <= window
    return False


def match_out_to_in_and_backfill(inter: pd.DataFrame, amount_tolerance: int, time_unit: str, window: int) -> pd.DataFrame:
    inter = inter.copy()
    inter["ìƒê³„í›„ë³´"] = False
    looks_out = inter["ë¹„ê³ "].fillna("").str.contains(r"ì—ê²Œ\s*ì†¡ê¸ˆ")
    inter["ìˆ˜ì‹ í›„ë³´"] = inter["ê±°ë˜ë°©ë²•"].str.contains(IN_METHOD_PATTERN, na=False) & (~looks_out)

    for (owner, counter), g in inter.groupby(["ê³„ì¢Œì£¼","ìƒëŒ€ê³„ì¢Œì£¼"], dropna=False):
        if pd.isna(counter):
            continue
        out_g = g[(g["ë°©í–¥"] == "out") & (~g["ì œì™¸ë¹„ê³ "])].sort_values("ì¼ì‹œ")
        if out_g.empty:
            continue
        in_pool = inter[(inter["ê³„ì¢Œì£¼"] == counter) & (inter["ìˆ˜ì‹ í›„ë³´"]) & (~inter["ì œì™¸ë¹„ê³ "])].sort_values("ì¼ì‹œ").copy()
        if in_pool.empty:
            continue
        out_rows, in_rows = out_g.to_dict("records"), in_pool.to_dict("records")
        i = j = 0
        while i < len(out_rows) and j < len(in_rows):
            ro, ri = out_rows[i], in_rows[j]
            amt_ok = abs(abs(ro["ê±°ë˜ëŒ€ê¸ˆ"]) - abs(ri["ê±°ë˜ëŒ€ê¸ˆ"])) <= amount_tolerance
            time_ok = time_within_window(ro["ì¼ì‹œ"], ri["ì¼ì‹œ"], time_unit, window)
            if amt_ok and time_ok:
                inter.loc[inter["orig_index"].isin([ro["orig_index"], ri["orig_index"]]), "ìƒê³„í›„ë³´"] = True
                ri_idx = inter.index[inter["orig_index"] == ri["orig_index"]][0]
                if not bool(inter.at[ri_idx, "ì œì™¸ë¹„ê³ "]):
                    if pd.isna(inter.at[ri_idx, "ìƒëŒ€ê³„ì¢Œì£¼"]) or inter.at[ri_idx, "ìƒëŒ€ê³„ì¢Œì£¼"] == "":
                        inter.at[ri_idx, "ìƒëŒ€ê³„ì¢Œì£¼"] = owner
                    if pd.isna(inter.at[ri_idx, "ë°©í–¥"]) or inter.at[ri_idx, "ë°©í–¥"] == "":
                        inter.at[ri_idx, "ë°©í–¥"] = "in"
                i += 1
                j += 1
            else:
                if ro["ì¼ì‹œ"] <= ri["ì¼ì‹œ"]:
                    i += 1
                else:
                    j += 1
    return inter


# ----------------------- ì—‘ì…€ ë¹Œë“œ & ìŠ¤íƒ€ì¼ -----------------------
if OPENPYXL_AVAILABLE:
    from openpyxl import Workbook  # type: ignore

    def _style_owner_sheet(ws):
        header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        col_amount = header_map.get("ê±°ë˜ëŒ€ê¸ˆ")
        col_balance = header_map.get("ì”ì•¡")
        col_candidate = header_map.get("ìƒê³„í›„ë³´")

        if col_amount:
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=col_amount).number_format = "#,##0"
        if col_balance:
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=col_balance).number_format = "#,##0"

        yellow = PatternFill(start_color="FFF9C0", end_color="FFF9C0", fill_type="solid")
        if col_candidate:
            for r in range(2, ws.max_row + 1):
                val = ws.cell(row=r, column=col_candidate).value
                if str(val).strip().lower() in ("true", "1"):
                    for c in range(1, ws.max_column + 1):
                        ws.cell(row=r, column=c).fill = yellow

        # í•©ê³„ í–‰
        last_row = ws.max_row
        total_row = last_row + 1
        ws.cell(row=total_row, column=1, value="í•©ê³„").font = Font(bold=True)
        if col_amount:
            L = get_column_letter(col_amount)
            ws.cell(row=total_row, column=col_amount, value=f"=SUM({L}2:{L}{last_row})").font = Font(bold=True)
            ws.cell(row=total_row, column=col_amount).number_format = "#,##0"
        if col_balance:
            L = get_column_letter(col_balance)
            ws.cell(row=total_row, column=col_balance, value=f"=SUM({L}2:{L}{last_row})").font = Font(bold=True)
            ws.cell(row=total_row, column=col_balance).number_format = "#,##0"

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    def _style_summary_sheet(ws):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=1, column=c).font = Font(bold=True)

        second_header_row = None
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == "ê³„ì¢Œì£¼" and ws.cell(row=r - 1, column=1).value is None:
                second_header_row = r
                break

        if second_header_row:
            for c in range(1, ws.max_column + 1):
                ws.cell(row=second_header_row, column=c).font = Font(bold=True)
            headers2 = {ws.cell(row=second_header_row, column=c).value: c for c in range(1, ws.max_column + 1)}
            for hdr in ["ìˆ˜ì‹ í•©ê³„", "ì†¡ê¸ˆí•©ê³„", "ìˆœì•¡(ìˆ˜ì‹ -ì†¡ê¸ˆ)"]:
                if hdr in headers2:
                    col = headers2[hdr]
                    rr = second_header_row + 1
                    while rr <= ws.max_row and ws.cell(row=rr, column=1).value is not None:
                        ws.cell(row=rr, column=col).number_format = "#,##0"
                        rr += 1

    def build_excel_bytes(inter_marked: pd.DataFrame) -> bytes:
        # 1) 1ì°¨ë¡œ ì—‘ì…€ ì‘ì„± (ê° ì‹œíŠ¸ + í†µí•© ìš”ì•½)
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            # ê³„ì¢Œì£¼ë³„ ì‹œíŠ¸
            for owner, sub in inter_marked.groupby("ê³„ì¢Œì£¼"):
                sub = sub.sort_values(["ìƒëŒ€ê³„ì¢Œì£¼", "ì¼ì‹œ"])
                cols = ["ì¼ì‹œ", "ê³„ì¢Œì£¼", "ê³„ì¢Œë²ˆí˜¸", "ìƒëŒ€ê³„ì¢Œì£¼", "ë°©í–¥", "ê±°ë˜ëŒ€ê¸ˆ", "ì”ì•¡", "ê±°ë˜ë°©ë²•", "ë¹„ê³ ", "ìƒê³„í›„ë³´"]
                sub[cols].to_excel(writer, index=False, sheet_name=str(owner))

            # í†µí•© ìš”ì•½ ì‹œíŠ¸ ì‘ì„±ìš© ë°ì´í„°
            m = inter_marked[inter_marked["ìƒê³„í›„ë³´"]].copy()
            pivot_sum = m.pivot_table(
                index=["ê³„ì¢Œì£¼", "ìƒëŒ€ê³„ì¢Œì£¼"],
                columns="ë°©í–¥",
                values="ê±°ë˜ëŒ€ê¸ˆ",
                aggfunc="sum",
                fill_value=0.0
            ).reset_index()
            if "in" not in pivot_sum.columns:
                pivot_sum["in"] = 0.0
            if "out" not in pivot_sum.columns:
                pivot_sum["out"] = 0.0
            pivot_sum["ìˆœì•¡(ìˆ˜ì‹ -ì†¡ê¸ˆ)"] = pivot_sum["in"] - pivot_sum["out"]

            pivot_cnt = m.groupby(["ê³„ì¢Œì£¼", "ìƒëŒ€ê³„ì¢Œì£¼", "ë°©í–¥"])["ê±°ë˜ëŒ€ê¸ˆ"].count().reset_index(name="ê±´ìˆ˜")
            pivot_cnt = pivot_cnt.pivot_table(index=["ê³„ì¢Œì£¼", "ìƒëŒ€ê³„ì¢Œì£¼"], columns="ë°©í–¥", values="ê±´ìˆ˜", fill_value=0).reset_index()
            pivot_cnt = pivot_cnt.rename(columns={"in": "ìˆ˜ì‹ ê±´ìˆ˜", "out": "ì†¡ê±´ìˆ˜"})

            merged_pairs = pd.merge(pivot_sum, pivot_cnt, on=["ê³„ì¢Œì£¼", "ìƒëŒ€ê³„ì¢Œì£¼"], how="left")
            merged_pairs = merged_pairs.rename(columns={"in": "ìˆ˜ì‹ í•©ê³„", "out": "ì†¡ê¸ˆí•©ê³„"})

            owner_sum = m.pivot_table(
                index=["ê³„ì¢Œì£¼"],
                columns="ë°©í–¥",
                values="ê±°ë˜ëŒ€ê¸ˆ",
                aggfunc="sum",
                fill_value=0.0
            ).reset_index()
            if "in" not in owner_sum.columns:
                owner_sum["in"] = 0.0
            if "out" not in owner_sum.columns:
                owner_sum["out"] = 0.0
            owner_sum["ìˆœì•¡(ìˆ˜ì‹ -ì†¡ê¸ˆ)"] = owner_sum["in"] - owner_sum["out"]
            owner_sum = owner_sum.rename(columns={"in": "ìˆ˜ì‹ í•©ê³„", "out": "ì†¡ê¸ˆí•©ê³„"})

            merged_pairs.to_excel(writer, index=False, sheet_name="í†µí•©", startrow=0)
            owner_sum.to_excel(writer, index=False, sheet_name="í†µí•©", startrow=len(merged_pairs) + 3)

        # 2) ìŠ¤íƒ€ì¼ ì ìš© (openpyxlë¡œ ì¬ì˜¤í”ˆ)
        buffer.seek(0)
        wb = load_workbook(buffer)
        for sheet_name in wb.sheetnames:
            if sheet_name == "í†µí•©":
                continue
            _style_owner_sheet(wb[sheet_name])
        _style_summary_sheet(wb["í†µí•©"])

        # 3) ë‹¤ì‹œ BytesIOë¡œ ì €ì¥
        out_buf = io.BytesIO()
        wb.save(out_buf)
        out_buf.seek(0)
        return out_buf.getvalue()
else:
    # openpyxlì´ ì—†ìœ¼ë©´ ìŠ¤íƒ€ì¼ ì—†ì´ xlsxwriterë¡œ ì €ì¥ë§Œ ìˆ˜í–‰
    def build_excel_bytes(inter_marked: pd.DataFrame) -> bytes:
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="xlsxwriter") as writer:
            for owner, sub in inter_marked.groupby("ê³„ì¢Œì£¼"):
                sub = sub.sort_values(["ìƒëŒ€ê³„ì¢Œì£¼", "ì¼ì‹œ"])
                cols = ["ì¼ì‹œ", "ê³„ì¢Œì£¼", "ê³„ì¢Œë²ˆí˜¸", "ìƒëŒ€ê³„ì¢Œì£¼", "ë°©í–¥", "ê±°ë˜ëŒ€ê¸ˆ", "ì”ì•¡", "ê±°ë˜ë°©ë²•", "ë¹„ê³ ", "ìƒê³„í›„ë³´"]
                sub[cols].to_excel(writer, index=False, sheet_name=str(owner))

            m = inter_marked[inter_marked["ìƒê³„í›„ë³´"]].copy()
            pivot_sum = m.pivot_table(
                index=["ê³„ì¢Œì£¼", "ìƒëŒ€ê³„ì¢Œì£¼"],
                columns="ë°©í–¥",
                values="ê±°ë˜ëŒ€ê¸ˆ",
                aggfunc="sum",
                fill_value=0.0
            ).reset_index()
            if "in" not in pivot_sum.columns:
                pivot_sum["in"] = 0.0
            if "out" not in pivot_sum.columns:
                pivot_sum["out"] = 0.0
            pivot_sum["ìˆœì•¡(ìˆ˜ì‹ -ì†¡ê¸ˆ)"] = pivot_sum["in"] - pivot_sum["out"]

            pivot_cnt = m.groupby(["ê³„ì¢Œì£¼", "ìƒëŒ€ê³„ì¢Œì£¼", "ë°©í–¥"])["ê±°ë˜ëŒ€ê¸ˆ"].count().reset_index(name="ê±´ìˆ˜")
            pivot_cnt = pivot_cnt.pivot_table(index=["ê³„ì¢Œì£¼", "ìƒëŒ€ê³„ì¢Œì£¼"], columns="ë°©í–¥", values="ê±´ìˆ˜", fill_value=0).reset_index()
            pivot_cnt = pivot_cnt.rename(columns={"in": "ìˆ˜ì‹ ê±´ìˆ˜", "out": "ì†¡ê±´ìˆ˜"})

            merged_pairs = pd.merge(pivot_sum, pivot_cnt, on=["ê³„ì¢Œì£¼", "ìƒëŒ€ê³„ì¢Œì£¼"], how="left")
            merged_pairs = merged_pairs.rename(columns={"in": "ìˆ˜ì‹ í•©ê³„", "out": "ì†¡ê¸ˆí•©ê³„"})

            owner_sum = m.pivot_table(index=["ê³„ì¢Œì£¼"], columns="ë°©í–¥", values="ê±°ë˜ëŒ€ê¸ˆ", aggfunc="sum", fill_value=0.0).reset_index()
            if "in" not in owner_sum.columns:
                owner_sum["in"] = 0.0
            if "out" not in owner_sum.columns:
                owner_sum["out"] = 0.0
            owner_sum["ìˆœì•¡(ìˆ˜ì‹ -ì†¡ê¸ˆ)"] = owner_sum["in"] - owner_sum["out"]
            owner_sum = owner_sum.rename(columns={"in": "ìˆ˜ì‹ í•©ê³„", "out": "ì†¡ê¸ˆí•©ê³„"})

            merged_pairs.to_excel(writer, index=False, sheet_name="í†µí•©", startrow=0)
            owner_sum.to_excel(writer, index=False, sheet_name="í†µí•©", startrow=len(merged_pairs) + 3)

        buffer.seek(0)
        return buffer.read()


# --------------------------- Streamlit UI ---------------------------
st.set_page_config(page_title="ê³„ì¢Œì£¼ê°„ ê±°ë˜ ë¶„ì„", page_icon="ğŸ’°", layout="wide")
st.title("ğŸ’° ê³„ì¢Œì£¼ê°„ ê±°ë˜ ë¶„ì„ ë„êµ¬ (Streamlit)")

with st.sidebar:
    st.subheader("ì„¤ì •")
    amount_tolerance = st.number_input("ê¸ˆì•¡ í—ˆìš© ì˜¤ì°¨ (ì›)", value=DEFAULT_AMOUNT_TOLERANCE, step=10_000, min_value=0)
    time_unit = st.selectbox("ì‹œê°„ ë‹¨ìœ„", ["day", "month", "year"], index=["day","month","year"].index(DEFAULT_TIME_UNIT))
    window = st.number_input("ì‹œê°„ì°½ í¬ê¸°", value=DEFAULT_WINDOW, step=1, min_value=1)
    st.caption("'ê¸‰ì—¬/ì›”ê¸‰/ë´‰ê¸‰/ê¸‰ì—¬ì…ê¸ˆ/salary'ê°€ ë“¤ì–´ê°„ ë¹„ê³ ëŠ” ìë™ ì œì™¸ë©ë‹ˆë‹¤.")

st.markdown("""
**ì‚¬ìš© ë°©ë²•**
1) ì•„ë˜ì— Excel íŒŒì¼(xlsx ê¶Œì¥, xls ê°€ëŠ¥)ì„ ì—…ë¡œë“œí•©ë‹ˆë‹¤.  
2) ì¢Œì¸¡ ì‚¬ì´ë“œë°”ì—ì„œ íŒŒë¼ë¯¸í„°ë¥¼ ì„¤ì •í•©ë‹ˆë‹¤.  
3) **ë¶„ì„ ì‹¤í–‰** ë²„íŠ¼ì„ ëˆ„ë¥´ë©´ ê²°ê³¼ë¥¼ ìƒì„±í•©ë‹ˆë‹¤.  
4) **ë‹¤ìš´ë¡œë“œ** ë²„íŠ¼ìœ¼ë¡œ ê²°ê³¼ xlsxë¥¼ ì €ì¥í•˜ì„¸ìš”.
""")

uploaded = st.file_uploader("ê±°ë˜ë‚´ì—­ Excel íŒŒì¼ ì—…ë¡œë“œ (xlsx ê¶Œì¥)", type=["xlsx", "xls", "xlsm"], accept_multiple_files=False)

if uploaded is not None:
    st.success(f"ì—…ë¡œë“œë¨: {uploaded.name}")
    with st.expander("ë¯¸ë¦¬ë³´ê¸° (ìƒìœ„ 10í–‰)", expanded=False):
        try:
            preview_df = _read_excel_any(uploaded, uploaded.name).head(10)
            st.dataframe(preview_df, use_container_width=True)
        except Exception as e:
            st.warning(str(e))

run = st.button("ğŸ” ë¶„ì„ ì‹¤í–‰", type="primary", use_container_width=True)

if run:
    if uploaded is None:
        st.error("ë¨¼ì € íŒŒì¼ì„ ì—…ë¡œë“œí•´ ì£¼ì„¸ìš”.")
    else:
        with st.spinner("ë¶„ì„ ì¤‘ì…ë‹ˆë‹¤..."):
            try:
                # 1) ë°ì´í„° ì½ê¸° & ì „ì²˜ë¦¬
                df = read_data_from_upload(uploaded, uploaded.name)
                inter = extract_inter_owner_transactions(df)
                inter_marked = match_out_to_in_and_backfill(
                    inter, amount_tolerance=amount_tolerance, time_unit=time_unit, window=window
                )

                # 2) í†µí•© ìš”ì•½ ì¼ë¶€ë¥¼ í™”ë©´ì—ì„œë„ ë³´ì—¬ì£¼ê¸°
                m = inter_marked[inter_marked["ìƒê³„í›„ë³´"]].copy()
                if m.empty:
                    st.info("ìƒê³„ í›„ë³´ë¡œ ë§¤ì¹­ëœ ê±°ë˜ê°€ ì—†ìŠµë‹ˆë‹¤. íŒŒë¼ë¯¸í„°ë¥¼ ì¡°ì •í•´ ë³´ì„¸ìš”.")
                else:
                    st.subheader("ìš”ì•½ ë¯¸ë¦¬ë³´ê¸°")
                    owner_sum = (
                        m.pivot_table(index=["ê³„ì¢Œì£¼"], columns="ë°©í–¥", values="ê±°ë˜ëŒ€ê¸ˆ", aggfunc="sum", fill_value=0.0)
                        .reset_index()
                        .rename(columns={"in": "ìˆ˜ì‹ í•©ê³„", "out": "ì†¡ê¸ˆí•©ê³„"})
                    )
                    owner_sum["ìˆœì•¡(ìˆ˜ì‹ -ì†¡ê¸ˆ)"] = owner_sum.get("ìˆ˜ì‹ í•©ê³„", 0.0) - owner_sum.get("ì†¡ê¸ˆí•©ê³„", 0.0)
                    st.dataframe(owner_sum, use_container_width=True)

                # 3) ì—‘ì…€ Bytes ìƒì„± & ë‹¤ìš´ë¡œë“œ ë²„íŠ¼
                xlsx_bytes = build_excel_bytes(inter_marked)
                label = "ğŸ“¥ ê²°ê³¼ íŒŒì¼ ë‹¤ìš´ë¡œë“œ (ê³„ì¢Œì£¼ê°„_ê±°ë˜_ë¶„ì„.xlsx)"
                if not OPENPYXL_AVAILABLE:
                    st.info("openpyxl ë¯¸ì„¤ì¹˜ë¡œ ì¸í•´ ì…€ ìŠ¤íƒ€ì¼ ì ìš© ì—†ì´ ì €ì¥ë©ë‹ˆë‹¤. 'pip install openpyxl' ì„¤ì¹˜ ì‹œ ìƒ‰ìƒ/í•©ê³„ ì„œì‹ì´ ì ìš©ë©ë‹ˆë‹¤.")
                st.download_button(
                    label=label,
                    data=xlsx_bytes,
                    file_name="ê³„ì¢Œì£¼ê°„_ê±°ë˜_ë¶„ì„.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

                st.success("ë¶„ì„ì´ ì™„ë£Œë˜ì—ˆìŠµë‹ˆë‹¤.")

            except Exception as e:
                st.error(f"ì˜¤ë¥˜ ë°œìƒ: {e}")
