# -*- coding: utf-8 -*-
# See in previous cell explanation (complete Streamlit app)
import io
import re
import difflib
import requests
import numpy as np
import pandas as pd
import streamlit as st
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="계좌주간 거래 분석(Excel+GitHub)", page_icon="🧮", layout="wide")
st.title("🧮 계좌주간 거래 분석 도구")

with st.sidebar:
    st.header("설정")
    st.caption("상계(왕복) 판정 및 순유출 의심 기준을 조정할 수 있어요.")
    WINDOW_DAYS = st.number_input("기간 허용(일)", min_value=0, max_value=60, value=7, step=1)
    AMOUNT_TOL_RATIO = st.number_input("금액 허용비율(%)", min_value=0.0, max_value=10.0, value=2.0, step=0.1) / 100.0
    AMOUNT_TOL_ABS = st.number_input("금액 절대허용(원)", min_value=0, max_value=10_000_000, value=10_000, step=1_000)
    GIFT_NET_THRESHOLD = st.number_input("증여의심(순유출 기준, 원)", min_value=0, max_value=10_000_000_000, value=10_000_000, step=1_000_000)
    st.markdown("---")
    st.subheader("GitHub에서 가져오기")
    repo_url = st.text_input(
        "저장소 URL",
        value="https://github.com/serene-gith/Account-transaction-history-analysis",
        placeholder="https://github.com/user/repo"
    )
    branch = st.text_input("브랜치", value="main")
    folder_path = st.text_input("폴더 경로(루트면 비움)", value="")
    gh_token = st.text_input("토큰(비공개 저장소)", type="password")
    st.caption("공개 저장소는 토큰 없이 동작합니다.")

st.markdown("#### 1) 파일 업로드 또는 2) GitHub에서 선택")

tab1, tab2 = st.tabs(["📤 로컬 파일 업로드", "🐙 GitHub에서 선택"])

uploaded_bytes = None
uploaded_name = None

with tab1:
    up = st.file_uploader("거래내역 Excel 파일 업로드 (.xlsx/.xlsm/.xls)", type=["xlsx","xlsm","xls"])
    if up is not None:
        uploaded_bytes = up.read()
        uploaded_name = up.name
        st.session_state["uploaded_bytes"] = uploaded_bytes
        st.session_state["uploaded_name"] = uploaded_name
        st.success(f"업로드됨: {uploaded_name}")
        # 미리보기
        try:
            preview = pd.read_excel(io.BytesIO(uploaded_bytes), nrows=10, header=None)
            st.expander("미리보기(상위 10행, 헤더 탐지 전)").dataframe(preview)
        except Exception as e:
            st.warning(f"미리보기 실패: {e}")

@st.cache_data(show_spinner=False)
def list_repo_xlsx(repo_url: str, branch: str, folder_path: str, token: Optional[str]):
    parts = repo_url.strip("/").split("/")
    owner, repo = parts[-2], parts[-1]
    path = folder_path.strip("/")
    api_url = f"https://api.github.com/repos/{owner}/{repo}/contents/{path}" if path else f"https://api.github.com/repos/{owner}/{repo}/contents"
    params = {"ref": branch}
    headers = {"Accept": "application/vnd.github+json"}
    if token:
        headers["Authorization"] = f"token {token}"
    r = requests.get(api_url, params=params, headers=headers, timeout=30)
    r.raise_for_status()
    items = r.json()
    if not isinstance(items, list):
        items = [items]
    excel_items = []
    for it in items:
        if it.get("type") == "file":
            name = it.get("name","")
            if name.lower().endswith((".xlsx",".xlsm",".xls")):
                excel_items.append({"name": name, "path": it.get("path")})
    excel_items.sort(key=lambda x: x["name"].lower())
    return owner, repo, excel_items

@st.cache_data(show_spinner=False)
def fetch_file_bytes(owner: str, repo: str, branch: str, file_path: str, token: Optional[str]):
    raw_url = f"https://raw.githubusercontent.com/{owner}/{repo}/{branch}/{file_path}"
    headers = {}
    if token:
        headers["Authorization"] = f"token {token}"
    r = requests.get(raw_url, headers=headers, timeout=60)
    if r.status_code == 404:
        api_url = f"https://api.github.com/repos/{owner}/{repo}/contents/{file_path}"
        params = {"ref": branch}
        api_headers = {"Accept": "application/vnd.github+json"}
        if token:
            api_headers["Authorization"] = f"token {token}"
        res = requests.get(api_url, params=params, headers=api_headers, timeout=60)
        res.raise_for_status()
        data = res.json()
        download_url = data.get("download_url")
        if not download_url:
            raise RuntimeError("download_url을 가져오지 못했습니다.")
        r = requests.get(download_url, timeout=60)
    r.raise_for_status()
    return r.content

with tab2:
    if st.button("🔍 목록 불러오기", use_container_width=True) and repo_url.strip():
        try:
            owner, repo, excel_list = list_repo_xlsx(repo_url, branch, folder_path, gh_token or None)
            st.session_state["gh_ctx"] = (owner, repo, excel_list)
        except Exception as e:
            st.error(f"GitHub 목록 오류: {e}")

    if "gh_ctx" in st.session_state:
        owner, repo, excel_list = st.session_state["gh_ctx"]
        if not excel_list:
            st.warning("엑셀 파일이 없습니다. 브랜치/폴더 경로를 확인하세요.")
        else:
            names = [x["name"] for x in excel_list]
            pick = st.selectbox("파일 선택", names, index=0)
            if st.button("⬇️ 이 파일 가져오기", type="primary"):
                try:
                    item = next(x for x in excel_list if x["name"] == pick)
                    data = fetch_file_bytes(owner, repo, branch, item["path"], gh_token or None)
                    uploaded_bytes = data
                    uploaded_name = pick
                    # ✅ 세션에 저장 (다른 버튼에서도 접근 가능)
                    st.session_state["uploaded_bytes"] = data
                    st.session_state["uploaded_name"] = pick

                    st.success(f"가져옴: {pick}")
                    prev = pd.read_excel(io.BytesIO(uploaded_bytes), nrows=10, header=None)
                    st.expander("미리보기(상위 10행, 헤더 탐지 전)").dataframe(prev)
                except Exception as e:
                    st.error(f"가져오기 오류: {e}")


SCAN_HEADER_ROWS = 20

SYN = {
    "date": ["거래일","거래일자","일자","date","transdate","날짜"],
    "owner": ["계좌주","예금주","소유자","owner","accountowner"],
    "owner_acct": ["계좌번호","계좌","account","owneraccount","fromaccount","출금계좌","출금계좌번호"],
    "cp_holder": ["상대계좌주","상대방","거래상대","수취인","받는이","받는사람","입금계좌주","counterparty"],
    "cp_acct": ["상대계좌번호","상대계좌","counterpartyaccount","입금계좌","입금계좌번호"],
    "desc": ["거래내역","적요","메모","내용","비고","summary","description","내역"],
    "amount": ["금액","거래금액","이체금액","amount","amt","금액(원)"],
    "credit": ["입금","입금액","수취금액","받은금액","credit","cr"],
    "debit": ["출금","출금액","송금액","보낸금액","debit","dr"],
    "from_holder": ["출금계좌주","보낸사람","송금인","fromholder","출금계좌명"],
    "from_acct": ["출금계좌","출금계좌번호","fromaccount","보낸계좌"],
    "to_holder": ["입금계좌주","받는사람","수취인명","toholder","입금계좌명"],
    "to_acct": ["입금계좌","입금계좌번호","toaccount","받는계좌"],
}

def _clean(s: str) -> str:
    s = str(s)
    s = s.replace("\n","").replace("\r","")
    s = re.sub(r"\s+", "", s)
    s = s.replace("(", "").replace(")", "")
    return s.lower()

def _to_number(x):
    if pd.isna(x):
        return np.nan
    s = str(x).strip()
    if re.match(r"^\(.*\)$", s):
        s = "-" + s[1:-1]
    s = s.replace(",", "")
    try:
        return float(s)
    except:
        return pd.to_numeric(s, errors="coerce")

def _best_match(col: str, candidates: List[str]) -> Optional[str]:
    if col in candidates:
        return col
    for c in candidates:
        if c in col or col in c:
            return c
    m = difflib.get_close_matches(col, candidates, n=1, cutoff=0.72)
    return m[0] if m else None

def _find_header(df: pd.DataFrame) -> pd.DataFrame:
    if not df.empty and all(isinstance(c, str) for c in df.columns):
        return df
    best_row, best_score = None, -1
    for r in range(min(SCAN_HEADER_ROWS, len(df))):
        row_vals = df.iloc[r].astype(str).tolist()
        cleaned = [_clean(v) for v in row_vals]
        tokens = sum(1 for v in cleaned if _best_match(v, sum(SYN.values(), [])) is not None)
        if tokens > best_score:
            best_score, best_row = tokens, r
    if best_row is None:
        return df
    new_cols = [_clean(x) for x in df.iloc[best_row].tolist()]
    df2 = df.iloc[best_row+1:].copy()
    df2.columns = new_cols
    df2.reset_index(drop=True, inplace=True)
    return df2

def _map_columns(cols: List[str]) -> Dict[str, Optional[str]]:
    def pick(keys: List[str]) -> Optional[str]:
        for key in keys:
            cands = [_clean(x) for x in SYN.get(key, [])]
            for col in cols:
                bm = _best_match(col, cands)
                if bm:
                    return col
        return None
    return {
        "owner": pick(["owner"]),
        "owner_acct": pick(["owner_acct"]),
        "cp_holder": pick(["cp_holder"]),
        "cp_acct": pick(["cp_acct"]),
        "amount": pick(["amount"]),
        "desc": pick(["desc"]),
        "date": pick(["date"]),
        "from_holder": pick(["from_holder"]),
        "from_acct": pick(["from_acct"]),
        "to_holder": pick(["to_holder"]),
        "to_acct": pick(["to_acct"]),
        "credit": pick(["credit"]),
        "debit": pick(["debit"]),
    }

def load_and_standardize_from_bytes(b: bytes) -> Tuple[pd.DataFrame, Dict[str,str], List[str]]:
    raw = pd.read_excel(io.BytesIO(b), header=None)
    df = _find_header(raw.copy())
    cols = [_clean(c) for c in df.columns]
    mapping = _map_columns(cols)

    has_single_amount = mapping["amount"] in cols if mapping["amount"] else False
    has_split_amounts = (mapping["credit"] in cols if mapping["credit"] else False) or (mapping["debit"] in cols if mapping["debit"] else False)

    if all(mapping.get(k) for k in ["from_holder","from_acct","to_holder","to_acct"]) and (has_single_amount or has_split_amounts):
        df_std = pd.DataFrame({
            "date": pd.to_datetime(df[mapping["date"]], errors="coerce") if mapping["date"] else pd.NaT,
            "from_holder": df[mapping["from_holder"]].astype(str).str.strip(),
            "from_acct": df[mapping["from_acct"]].astype(str).str.strip(),
            "to_holder": df[mapping["to_holder"]].astype(str).str.strip(),
            "to_acct": df[mapping["to_acct"]].astype(str).str.strip(),
            "desc": df[mapping["desc"]].astype(str) if mapping["desc"] else "",
        })
        if has_single_amount:
            df_std["amount"] = df[mapping["amount"]].apply(_to_number)
            amount_info = f"amount ← {mapping['amount']}"
        else:
            cr = df[mapping["credit"]].apply(_to_number) if mapping["credit"] else 0.0
            dr = df[mapping["debit"]].apply(_to_number) if mapping["debit"] else 0.0
            df_std["amount"] = cr.where(cr.notna() & (cr != 0), 0) - dr.where(dr.notna() & (dr != 0), 0)
            amount_info = f"amount ← +{mapping.get('credit')} -{mapping.get('debit')}"
        schema = "B(from/to)"
    elif all(mapping.get(k) for k in ["owner","owner_acct","cp_holder","cp_acct"]) and (has_single_amount or has_split_amounts):
        owner = df[mapping["owner"]].astype(str).str.strip()
        owner_acct = df[mapping["owner_acct"]].astype(str).str.strip()
        cp_holder = df[mapping["cp_holder"]].astype(str).str.strip()
        cp_acct = df[mapping["cp_acct"]].astype(str).str.strip()
        if has_single_amount:
            amt = df[mapping["amount"]].apply(_to_number)
            amount_info = f"amount ← {mapping['amount']}"
            from_holder, to_holder = owner, cp_holder
            from_acct, to_acct = owner_acct, cp_acct
        else:
            cr = df[mapping["credit"]].apply(_to_number) if mapping["credit"] else 0.0
            dr = df[mapping["debit"]].apply(_to_number) if mapping["debit"] else 0.0
            amt = cr.where(cr.notna() & (cr != 0), 0) - dr.where(dr.notna() & (dr != 0), 0)
            amount_info = f"amount ← +{mapping.get('credit')} -{mapping.get('debit')}"
            from_holder = np.where(amt >= 0, cp_holder, owner)
            to_holder = np.where(amt >= 0, owner, cp_holder)
            from_acct = np.where(amt >= 0, cp_acct, owner_acct)
            to_acct = np.where(amt >= 0, owner_acct, cp_acct)

        df_std = pd.DataFrame({
            "date": pd.to_datetime(df[mapping["date"]], errors="coerce") if mapping["date"] else pd.NaT,
            "from_holder": from_holder,
            "from_acct": from_acct,
            "to_holder": to_holder,
            "to_acct": to_acct,
            "amount": amt,
            "desc": df[mapping["desc"]].astype(str) if mapping["desc"] else "",
        })
        schema = "A(owner/counterparty)"
    elif all(k in df.columns for k in ["date","from_holder","from_acct","to_holder","to_acct","amount"]):
        df_std = pd.DataFrame({
            "date": pd.to_datetime(df["date"], errors="coerce"),
            "from_holder": df["from_holder"].astype(str).str.strip(),
            "from_acct": df["from_acct"].astype(str).str.strip(),
            "to_holder": df["to_holder"].astype(str).str.strip(),
            "to_acct": df["to_acct"].astype(str).str.strip(),
            "amount": pd.to_numeric(df["amount"], errors="coerce"),
            "desc": df["desc"].astype(str) if "desc" in df.columns else "",
        })
        amount_info = "already standardized"
        schema = "C(standardized)"
    else:
        raise ValueError(f"컬럼 자동 인식 실패. 감지된 컬럼: {cols}")

    df_std = df_std.dropna(subset=["from_holder","to_holder","amount"]).copy()
    if df_std["date"].notna().any():
        df_std.sort_values(["date"], inplace=True, kind="mergesort")
    df_std.reset_index(drop=True, inplace=True)

    mapping_summary = {
        "schema_detected": schema,
        "amount_info": amount_info,
        **mapping
    }
    return df_std, mapping_summary, cols

from dataclasses import dataclass

@dataclass
class Match:
    i: int
    j: int
    reason: str

def _amount_close(a: float, b: float, tol_abs: float, tol_ratio: float) -> bool:
    if pd.isna(a) or pd.isna(b):
        return False
    if abs(a - b) <= tol_abs:
        return True
    base = max(abs(a), abs(b), 1.0)
    return abs(a - b) / base <= tol_ratio

def find_matches(df: pd.DataFrame) -> List[Match]:
    matches: List[Match] = []
    used = set()
    df2 = df.copy()
    df2["pair_key"] = df2.apply(lambda r: frozenset([r["from_holder"], r["to_holder"]]), axis=1)
    for key, g in df2.groupby("pair_key", sort=False):
        if len(key) != 2:
            continue
        a, b = list(key)
        pos = g[(g["from_holder"] == a) & (g["to_holder"] == b)]
        neg = g[(g["from_holder"] == b) & (g["to_holder"] == a)]
        for i in pos.index:
            if i in used: 
                continue
            for j in neg.index:
                if j in used:
                    continue
                if not _amount_close(df.loc[i,"amount"], df.loc[j,"amount"], AMOUNT_TOL_ABS, AMOUNT_TOL_RATIO):
                    continue
                d1, d2 = df.loc[i,"date"], df.loc[j,"date"]
                if pd.notna(d1) and pd.notna(d2):
                    if abs((d2 - d1).days) > WINDOW_DAYS:
                        continue
                    reason = f"금액·기간 유사(≤{WINDOW_DAYS}일)"
                else:
                    reason = "금액 유사(날짜 없음)"
                used.update([i,j])
                matches.append(Match(i=i, j=j, reason=reason))
                break
    return matches

def build_owner_sheets(df: pd.DataFrame, matches: List[Match]):
    match_id = pd.Series(index=df.index, dtype="Int64")
    reason = pd.Series(index=df.index, dtype="string")
    for k, m in enumerate(matches, start=1):
        match_id.loc[m.i] = k
        match_id.loc[m.j] = k
        reason.loc[m.i] = m.reason
        reason.loc[m.j] = m.reason
    df2 = df.copy()
    df2["match_id"] = match_id
    df2["match_reason"] = reason
    df2["counterparty_holder"] = df2["to_holder"]
    df2["counterparty_account"] = df2["to_acct"]
    owners = sorted(set(df2["from_holder"]).union(set(df2["to_holder"])))
    sheets = {}
    base_cols = ["date","from_holder","from_acct","to_holder","to_acct","counterparty_holder","counterparty_account","amount","desc","match_id","match_reason"]
    for owner in owners:
        sub = df2[(df2["from_holder"] == owner) | (df2["to_holder"] == owner)].copy()
        sub = sub[sub["from_holder"] != sub["to_holder"]]
        if sub["date"].notna().any():
            sub = sub.sort_values(["date","amount"], kind="mergesort")
        else:
            sub = sub.sort_values(["amount"], kind="mergesort")
        present = [c for c in base_cols if c in sub.columns]
        sheets[owner] = sub[present].reset_index(drop=True)
    return sheets, df2

def make_summary(df: pd.DataFrame) -> pd.DataFrame:
    pairs = df.groupby(["from_holder", "to_holder"], dropna=False)["amount"].sum().reset_index()
    pairs = pairs[pairs["from_holder"] != pairs["to_holder"]].copy()
    pairs["pair_key"] = pairs.apply(lambda r: tuple(sorted([r["from_holder"], r["to_holder"]])), axis=1)
    recs = []
    for key, g in pairs.groupby("pair_key"):
        a, b = key
        a2b = g[(g["from_holder"] == a) & (g["to_holder"] == b)]["amount"].sum()
        b2a = g[(g["from_holder"] == b) & (g["to_holder"] == a)]["amount"].sum()
        net = a2b - b2a
        donor = a if net > 0 else b
        receiver = b if net > 0 else a
        suspicious = abs(net) >= GIFT_NET_THRESHOLD
        recs.append({
            "holder_A": a,
            "holder_B": b,
            "A->B_total": a2b,
            "B->A_total": b2a,
            "net_A_to_B": net,
            "possible_donor": donor if suspicious else "",
            "possible_receiver": receiver if suspicious else "",
            "gift_risk_hint": "순유출 과다" if suspicious else ""
        })
    result_df = pd.DataFrame(recs)
    if len(result_df) == 0:
        return result_df
    return result_df.sort_values("net_A_to_B", ascending=False, ignore_index=True)

def render_and_download(df: pd.DataFrame, sheets: Dict[str, pd.DataFrame], summary: pd.DataFrame, mapping_summary: Dict[str,str], detected_cols: List[str]) -> bytes:
    import io
    yellow = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        summary.to_excel(writer, index=False, sheet_name="Summary")
        for owner, df_owner in sheets.items():
            sheet_name = owner[:31] or "Unknown"
            df_owner.to_excel(writer, index=False, sheet_name=sheet_name)
        df.to_excel(writer, index=False, sheet_name="_Master")
        map_df = pd.DataFrame([mapping_summary])
        det_df = pd.DataFrame({"detected_columns": detected_cols})
        map_df.to_excel(writer, index=False, sheet_name="Mapping")
        det_df.to_excel(writer, index=False, sheet_name="Mapping", startrow=map_df.shape[0]+2)

    buffer.seek(0)
    wb = load_workbook(buffer)
    for ws in wb.worksheets:
        if ws.title == "Mapping":
            for col in range(1, ws.max_column + 1):
                max_len = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1))
                ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 60)
            continue
        if ws.title == "Summary":
            for col in range(1, ws.max_column + 1):
                max_len = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1))
                ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 40)
            continue
        headers = {cell.value: idx for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), start=1)}
        mid_col = headers.get("match_id")
        if mid_col:
            for row in ws.iter_rows(min_row=2):
                val = row[mid_col - 1].value
                if val not in (None, "", 0):
                    for cell in row:
                        cell.fill = yellow
        for col in range(1, ws.max_column + 1):
            max_len = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1))
            ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 40)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()

st.markdown("---")
run_col, info_col = st.columns([1,2])
with run_col:
    run = st.button("🔎 분석 실행", type="primary", use_container_width=True)
with info_col:
    st.info("파일을 업로드하거나 GitHub에서 선택한 뒤 **분석 실행**을 누르면 결과 XLSX가 생성됩니다.")

if run:
    # 세션에서 가져오기
    if "uploaded_bytes" in st.session_state:
        uploaded_bytes = st.session_state["uploaded_bytes"]
        uploaded_name = st.session_state["uploaded_name"]
    
    if not uploaded_bytes:
        st.error("분석할 파일이 없습니다. 업로드하거나 GitHub에서 선택하세요.")
        st.stop()
    try:
        df, mapping_summary, detected_cols = load_and_standardize_from_bytes(uploaded_bytes)
        matches = find_matches(df)
        sheets, master = build_owner_sheets(df, matches)
        summary = make_summary(df)
        result_bytes = render_and_download(master, sheets, summary, mapping_summary, detected_cols)
        st.success("분석 완료! 아래 버튼으로 결과 파일을 저장하세요.")
        fname = (uploaded_name or "결과") + "_분석결과.xlsx"
        st.download_button("💾 결과 XLSX 다운로드", data=result_bytes, file_name=fname, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        st.dataframe(summary, use_container_width=True, height=300)
    except Exception as e:
        st.exception(e)
        st.error("분석 중 오류가 발생했습니다. Mapping 시트를 참고해 컬럼명을 확인해 주세요.")